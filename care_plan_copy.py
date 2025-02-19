import streamlit as st
import pandas as pd
from datetime import datetime
from openai import OpenAI
import os
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO
import tempfile

# 環境変数の読み込み
load_dotenv()

# OpenAI クライアントの設定
client = OpenAI(
    api_key=os.getenv('OPENAI_API_KEY')
)

# ページ設定
st.set_page_config(
    page_title="EGAO-AI デモ",
    page_icon="👥",
    layout="wide"
)

# セッション状態の初期化
if 'generated_care_plan' not in st.session_state:
    st.session_state.generated_care_plan = None

def get_adl_status_color(status):
    """ADL状態に応じたカラーコードを返す"""
    colors = {
        "要全介助": "#ff6b6b",
        "一部介助": "#ffd93d",
        "見守り": "#a3dc2e",
        "自立": "#4CAF50"
    }
    return colors.get(status, "#ffffff")

def get_adl_description(item, status):
    """ADL項目と状態に応じた説明文を返す"""
    descriptions = {
        "食事": {
            "要全介助": "食事の全過程で介助が必要",
            "一部介助": "食事の一部で介助が必要",
            "見守り": "声かけ・見守りが必要",
            "自立": "自力で食事が可能"
        },
        "排泄": {
            "要全介助": "排泄の全過程で介助が必要",
            "一部介助": "排泄の一部で介助が必要",
            "見守り": "声かけ・見守りが必要",
            "自立": "自力で排泄が可能"
        },
        "入浴": {
            "要全介助": "入浴の全過程で介助が必要",
            "一部介助": "入浴の一部で介助が必要",
            "見守り": "声かけ・見守りが必要",
            "自立": "自力で入浴が可能"
        }
    }
    default_descriptions = {
        "要全介助": "常時介助が必要",
        "一部介助": "部分的な介助が必要",
        "見守り": "声かけ・見守りが必要",
        "自立": "自力で可能"
    }
    return descriptions.get(item, default_descriptions).get(status, "")

def create_care_plan_excel(user_info, adl_data, care_plan):
    """ケアプランをエクセルファイルとして生成"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ケアプラン"

        # スタイル設定
        title_font = Font(size=14, bold=True)
        header_font = Font(size=12, bold=True)
        
        # タイトルと基本情報
        ws['A1'] = "居宅サービス計画書（1）"
        ws['A1'].font = title_font
        
        ws['A3'] = "利用者基本情報"
        ws['A3'].font = header_font
        
        # 基本情報の記入
        basic_info = [
            ("利用者名", user_info.get('name', '')),
            ("年齢", f"{user_info.get('age', '')}歳"),
            ("性別", user_info.get('gender', '')),
            ("要介護度", user_info.get('care_level', '')),
            ("家族構成", user_info.get('family_structure', '')),
            ("キーパーソン", user_info.get('key_person', ''))
        ]
        
        for i, (label, value) in enumerate(basic_info, start=4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
        
        # ADLデータの記入
        ws['A10'] = "ADL評価"
        ws['A10'].font = header_font
        
        for i, (item, value) in enumerate(adl_data.items(), start=11):
            ws[f'A{i}'] = item
            ws[f'B{i}'] = value
        
        # ケアプラン内容の記入
        current_row = len(adl_data) + 13
        ws[f'A{current_row}'] = "ケアプラン"
        ws[f'A{current_row}'].font = header_font
        
        # ケアプランの内容を整形して記入
        plan_sections = care_plan.split('\n\n')
        for section in plan_sections:
            current_row += 1
            ws[f'A{current_row}'] = section
            ws.row_dimensions[current_row].height = 30
            
            # セルの書式設定
            cell = ws[f'A{current_row}']
            cell.alignment = Alignment(wrap_text=True)
        
        # 列幅の調整
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60

        # Excelファイルをバイトストリームとして保存
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer

    except Exception as e:
        st.error(f"エクセルファイル生成中にエラーが発生しました: {str(e)}")
        return None

def generate_care_plan(user_info, adl_data, client_needs):
    """OpenAI APIを使用してケアプラン生成"""
    try:
        # プロンプトの構築
        prompt = f"""
利用者基本情報:
- 名前: {user_info['name']}
- 年齢: {user_info['age']}歳
- 性別: {user_info['gender']}
- 要介護度: {user_info['care_level']}
- 家族構成: {user_info['family_structure']}
- キーパーソン: {user_info['key_person']}

ADL状況:
{pd.DataFrame([adl_data]).T.to_string()}

利用者の要望:
{client_needs}

上記の情報に基づいて、以下の項目を含む具体的なケアプランを作成してください：
1. 解決すべき課題
2. 長期目標（6ヶ月）
3. 短期目標（3ヶ月）
4. 具体的なサービス内容
"""

        # OpenAI APIを使用してケアプラン生成
        response = client.chat.completions.create(
            model="gpt-4-turbo-preview",
            messages=[
                {"role": "system", "content": "あなたは経験豊富な介護支援専門員です。"},
                {"role": "user", "content": prompt}
            ],
            temperature=0.7,
            max_tokens=2000
        )

        return response.choices[0].message.content

    except Exception as e:
        st.error(f"ケアプラン生成中にエラーが発生しました: {str(e)}")
        return None

def render_adl_input_section(items, category_name):
    """ADL入力セクションのレンダリング"""
    st.subheader(category_name)
    category_data = {}
    
    for item in items:
        status = st.selectbox(
            item,
            ["要全介助", "一部介助", "見守り", "自立"],
            key=f"adl_{item}"
        )
        category_data[item] = status
        
        # 説明文の表示
        description = get_adl_description(item, status)
        if description:
            st.markdown(f"<div style='color: {get_adl_status_color(status)}'>{description}</div>", unsafe_allow_html=True)
    
    return category_data

def main():
    st.title("EGAO-AI ケアプラン作成支援システム")
    
    # サイドバーと基本メニュー（既存の実装を維持）
    with st.sidebar:
        st.header("メニュー")
        page = st.radio(
            "選択してください",
            ["基本情報入力", "ADLデータ入力", "ケアプラン生成", "履歴管理"]
        )

    # 各ページの実装
    if page == "基本情報入力":
        st.header("基本情報入力")
        
        col1, col2 = st.columns(2)
        with col1:
            name = st.text_input("利用者名")
            age = st.number_input("年齢", min_value=0, max_value=150)
            gender = st.selectbox("性別", ["選択してください", "男性", "女性"])
            
        with col2:
            care_level = st.selectbox(
                "要介護度",
                ["選択してください", "要介護1", "要介護2", "要介護3", "要介護4", "要介護5"]
            )
            family_structure = st.text_input("家族構成")
            key_person = st.text_input("キーパーソン")
        
        if st.button("基本情報を保存", type="primary"):
            if name and age > 0 and gender != "選択してください" and care_level != "選択してください":
                st.session_state.user_info = {
                    "name": name,
                    "age": age,
                    "gender": gender,
                    "care_level": care_level,
                    "family_structure": family_structure,
                    "key_person": key_person
                }
                st.success("基本情報が保存されました")
                st.write(st.session_state.user_info)
            else:
                st.warning("必須項目を入力してください")

    elif page == "ADLデータ入力":
        st.header("ADLデータ入力")
        
        adl_categories = {
            "基本動作": ["食事", "排泄", "入浴", "移動", "着替え", "整容"],
            "認知・コミュニケーション": ["コミュニケーション", "認知機能", "睡眠"],
            "社会生活": ["服薬管理", "金銭管理", "買い物"]
        }
        
        tabs = st.tabs(list(adl_categories.keys()))
        all_adl_data = {}
        
        for tab, (category, items) in zip(tabs, adl_categories.items()):
            with tab:
                category_data = render_adl_input_section(items, category)
                all_adl_data.update(category_data)
        
        if st.button("ADLデータを保存", type="primary"):
            st.session_state.adl_data = all_adl_data
            st.success("ADLデータが保存されました")
            st.write(pd.DataFrame([all_adl_data]).T)
            
    elif page == "ケアプラン生成":
        st.header("ケアプラン生成")
        
        if 'user_info' not in st.session_state or 'adl_data' not in st.session_state:
            st.warning("基本情報とADLデータを先に入力してください")
            return
        
        st.subheader("利用者の要望")
        client_needs = st.text_area(
            "具体的な要望を入力してください",
            height=100,
            placeholder="例：母親の結婚式に参加したい"
        )
        
        if st.button("ケアプランを生成", type="primary"):
            if not client_needs:
                st.warning("利用者の要望を入力してください")
                return
                
            with st.spinner("ケアプランを生成中..."):
                care_plan = generate_care_plan(
                    st.session_state.user_info,
                    st.session_state.adl_data,
                    client_needs
                )
                
                if care_plan:
                    st.session_state.generated_care_plan = care_plan
                    st.success("ケアプランが生成されました")
                    
                    st.subheader("生成されたケアプラン")
                    st.markdown(care_plan)
                    
                    # ダウンロードボタンのセクション
                    col1, col2 = st.columns(2)
                    with col1:
                        st.download_button(
                            "テキスト形式でダウンロード",
                            care_plan,
                            "care_plan.txt",
                            "text/plain"
                        )
                    
                    with col2:
                        # エクセルファイルの生成とダウンロード
                        excel_buffer = create_care_plan_excel(
                            st.session_state.user_info,
                            st.session_state.adl_data,
                            care_plan
                        )
                        
                        if excel_buffer:
                            st.download_button(
                                "エクセル形式でダウンロード",
                                excel_buffer,
                                "care_plan.xlsx",
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
    
    elif page == "履歴管理":
        st.header("履歴管理")
        st.info("この機能は開発中です")

if __name__ == "__main__":
    main()