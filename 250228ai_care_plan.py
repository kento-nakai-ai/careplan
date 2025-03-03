# 必要なライブラリのインポート
import streamlit as st  # Webアプリケーションフレームワーク
import pandas as pd  # データ操作・分析用ライブラリ
from datetime import datetime  # 日付・時刻操作用
import os  # OS関連の操作用
from dotenv import load_dotenv  # 環境変数読み込み用
import openpyxl  # Excel操作用
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # Excelのスタイル設定用
from io import BytesIO  # バイナリデータ操作用
import tempfile  # 一時ファイル作成用
from reportlab.pdfgen import canvas  # PDF生成用
from reportlab.pdfbase import pdfmetrics  # PDFフォント管理用
from reportlab.pdfbase.ttfonts import TTFont  # PDFフォント設定用
from reportlab.lib.pagesizes import A4  # PDFページサイズ設定用
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer  # PDF要素作成用
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # PDFスタイル設定用
import zipfile  # ZIP圧縮用
from reportlab.pdfbase.cidfonts import UnicodeCIDFont  # 日本語フォント用
from reportlab.lib.units import mm  # PDFの単位設定用
import yaml  # YAML設定ファイル読み込み用
import traceback  # エラートレース用
import openai  # OpenAI APIクライアント
import json  # JSONモジュール
import shutil  # ファイルコピー用

# 環境変数の読み込み（ローカル開発用）
load_dotenv()

# OpenAI APIキーの設定
openai_api_key = st.secrets.get('OPENAI_API_KEY') or os.getenv('OPENAI_API_KEY')

if not openai_api_key:
    st.error('OpenAI APIキーが設定されていません。')
    st.stop()

# OpenAI APIクライアントの初期化
try:
    client = openai.OpenAI(api_key=openai_api_key)
except Exception as e:
    st.error(f"OpenAI APIクライアントの初期化に失敗しました: {str(e)}")
    st.stop()

# Streamlitページの基本設定
st.set_page_config(
    page_title="EGAO-AI デモ",  # ページタイトル
    page_icon="👥",  # ページアイコン
    layout="wide"  # ページレイアウト（ワイド）
)

# セッション状態の初期化
# 生成されたケアプランを保存する変数
if 'generated_care_plan' not in st.session_state:
    st.session_state.generated_care_plan = None

# ケアプラン履歴を保存するリスト
if 'care_plan_history' not in st.session_state:
    st.session_state.care_plan_history = []

def get_adl_status_color(status):
    """
    ADL状態に応じたカラーコードを返す関数
    
    Args:
        status (str): ADL状態（要全介助、一部介助、見守り、自立）
    
    Returns:
        str: カラーコード（HEX形式）
    """
    colors = {
        "要全介助": "#ff6b6b",  # 赤色 - 最も介助が必要な状態
        "一部介助": "#ffd93d",  # 黄色 - 部分的な介助が必要な状態
        "見守り": "#a3dc2e",    # 薄緑色 - 自立に近い状態
        "自立": "#4CAF50"       # 緑色 - 完全に自立している状態
    }
    return colors.get(status, "#ffffff")  # 該当するステータスがない場合は白色を返す

def get_adl_description(item, status):
    """
    ADL項目と状態に応じた説明文を返す関数
    
    Args:
        item (str): ADL項目（食事、排泄など）
        status (str): ADL状態（要全介助、一部介助、見守り、自立）
    
    Returns:
        str: 詳細な説明文
    """
    # 各ADL項目ごとの状態説明を辞書形式で定義
    descriptions = {
        "食事": {
            "要全介助": "食事の全過程で介助が必要（食事の準備から片付けまで、食べる動作すべてに介助が必要）",
            "一部介助": "食事の一部で介助が必要（食べ物を刻む、スプーンで掬うなどの補助が必要）",
            "見守り": "声かけ・見守りが必要（自力で食べられるが、誤嚥防止などの観察が必要）",
            "自立": "自力で食事が可能（準備から片付けまで完全に自立している）"
        },
        "排泄": {
            "要全介助": "排泄の全過程で介助が必要（トイレまでの移動、衣服の着脱、排泄後の処理すべてに介助が必要）",
            "一部介助": "排泄の一部で介助が必要（衣服の着脱の補助や、後始末の一部介助が必要）",
            "見守り": "声かけ・見守りが必要（自力で可能だが、安全確認のため見守りが必要）",
            "自立": "自力で排泄が可能（トイレまでの移動から後始末まで完全に自立している）"
        },
        "入浴": {
            "要全介助": "入浴の全過程で介助が必要（浴室への移動、衣服の着脱、洗体、洗髪すべてに介助が必要）",
            "一部介助": "入浴の一部で介助が必要（背中を洗う、髪を洗うなどの部分的な介助が必要）",
            "見守り": "声かけ・見守りが必要（自力で入浴可能だが、転倒防止のため見守りが必要）",
            "自立": "自力で入浴が可能（準備から後片付けまで完全に自立している）"
        },
        "移動": {
            "要全介助": "移動の全過程で介助が必要（ベッドから車椅子への移乗を含め、すべての移動に介助が必要）",
            "一部介助": "移動の一部で介助が必要（歩行器や杖を使用し、部分的な支援が必要）",
            "見守り": "声かけ・見守りが必要（自力で移動可能だが、安全確認のため見守りが必要）",
            "自立": "自力で移動が可能（補助具の使用の有無に関わらず、安全に移動できる）"
        },
        "着替え": {
            "要全介助": "着替えの全過程で介助が必要（衣服の選択から着脱まですべてに介助が必要）",
            "一部介助": "着替えの一部で介助が必要（ボタンの掛け外しや、靴下の着脱などの補助が必要）",
            "見守り": "声かけ・見守りが必要（自力で着替え可能だが、確認や声かけが必要）",
            "自立": "自力で着替えが可能（衣服の選択から着脱まで完全に自立している）"
        },
        "整容": {
            "要全介助": "整容の全過程で介助が必要（洗顔、歯磨き、整髪すべてに介助が必要）",
            "一部介助": "整容の一部で介助が必要（髭剃りや整髪などの部分的な介助が必要）",
            "見守り": "声かけ・見守りが必要（自力で整容可能だが、確認や声かけが必要）",
            "自立": "自力で整容が可能（身だしなみを整えることが完全に自立している）"
        },
        "コミュニケーション": {
            "要全介助": "コミュニケーションが極めて困難（意思疎通がほとんど取れない）",
            "一部介助": "基本的な意思疎通に支援が必要（簡単な言葉や身振りでの意思疎通が可能）",
            "見守り": "時々支援が必要（複雑な会話に困難があるが、基本的な意思疎通は可能）",
            "自立": "円滑なコミュニケーションが可能（会話に問題なく参加できる）"
        },
        "認知機能": {
            "要全介助": "重度の認知機能低下（日常生活のほとんどの判断に支援が必要）",
            "一部介助": "中等度の認知機能低下（日常生活の一部の判断に支援が必要）",
            "見守り": "軽度の認知機能低下（時々判断に迷いがあるが、声かけで対応可能）",
            "自立": "認知機能は良好（日常生活の判断に問題なし）"
        },
        "睡眠": {
            "要全介助": "睡眠のリズムが大きく乱れている（昼夜逆転や不眠が顕著）",
            "一部介助": "睡眠に問題があり支援が必要（入眠困難や中途覚醒がある）",
            "見守り": "時々睡眠に乱れがある（軽度の不眠や早朝覚醒がある）",
            "自立": "良好な睡眠が取れている（睡眠のリズムが安定している）"
        },
        "服薬管理": {
            "要全介助": "服薬管理が全くできない（薬の準備から服用まですべてに介助が必要）",
            "一部介助": "服薬管理に部分的な支援が必要（薬の仕分けや時間の管理に援助が必要）",
            "見守り": "声かけがあれば服薬管理可能（確認や促しが必要）",
            "自立": "自力で服薬管理が可能（処方薬の管理を適切に行える）"
        },
        "金銭管理": {
            "要全介助": "金銭管理が全くできない（収支の理解や管理がまったくできない）",
            "一部介助": "基本的な金銭管理に支援が必要（日々の出納の一部に支援が必要）",
            "見守り": "声かけがあれば金銭管理可能（確認や助言が必要）",
            "自立": "自力で金銭管理が可能（収支の把握や管理が適切にできる）"
        },
        "買い物": {
            "要全介助": "買い物が全くできない（商品の選択から支払いまですべてに介助が必要）",
            "一部介助": "買い物に部分的な支援が必要（商品の選択や支払いの一部に援助が必要）",
            "見守り": "声かけがあれば買い物可能（確認や助言が必要）",
            "自立": "自力で買い物が可能（必要な物品の購入を適切に行える）"
        }
    }
    default_descriptions = {
        "要全介助": "常時介助が必要（自力での実施が困難）",
        "一部介助": "部分的な介助が必要（一部自力で行えるが支援が必要）",
        "見守り": "声かけ・見守りが必要（安全確認のため観察が必要）",
        "自立": "自力で可能（支援なしで実施できる）"
    }
    return descriptions.get(item, default_descriptions).get(status, "")

def create_care_plan_excel(user_info, adl_data, care_plan):
    """ケアプランをcare_plan_1.xlsの形式で生成"""
    try:
        wb = openpyxl.Workbook()
        
        # 第1表シート
        ws1 = wb.active
        ws1.title = "第1表"
        
        # タイトルと作成日
        ws1.merge_cells('A1:J1')
        ws1["A1"] = "居宅サービス計画書（１）"
        ws1.merge_cells('K1:M1')
        ws1["K1"] = f"作成年月日：{datetime.now().strftime('%Y年%m月%d日')}"
        
        # 計画区分チェックボックス
        ws1.merge_cells('A3:B3')
        plan_status = []
        if user_info['plan_status']['initial']:
            plan_status.append("☑初回")
        else:
            plan_status.append("□初回")
        if user_info['plan_status']['introduced']:
            plan_status.append("☑紹介")
        else:
            plan_status.append("□紹介")
        if user_info['plan_status']['continuous']:
            plan_status.append("☑継続")
        else:
            plan_status.append("□継続")
        ws1["A3"] = " ・ ".join(plan_status)
        
        # 認定状況チェックボックス
        ws1.merge_cells('K3:M3')
        cert_status = []
        if user_info['plan_status']['certified']:
            cert_status.append("☑認定済")
        else:
            cert_status.append("□認定済")
        if user_info['plan_status']['applying']:
            cert_status.append("☑申請中")
        else:
            cert_status.append("□申請中")
        ws1["K3"] = " ・ ".join(cert_status)
        
        # 利用者基本情報
        ws1["A5"] = "利用者名"
        ws1.merge_cells('B5:D5')
        ws1["B5"] = f"{user_info['name']}様"
        ws1.merge_cells('E5:M5')
        ws1["E5"] = f"生年月日 {user_info['birth_date']}"
        
        ws1["A6"] = "住所"
        ws1.merge_cells('B6:M6')
        ws1["B6"] = user_info['address']
        
        ws1["A7"] = "居宅サービス計画作成者氏名"
        ws1.merge_cells('B7:M7')
        ws1["B7"] = user_info['care_manager']
        
        ws1["A8"] = "居宅介護支援事業者・事業所名及び所在地"
        ws1.merge_cells('B8:M8')
        ws1["B8"] = f"{user_info['care_office']}　{user_info['office_address']}"
        
        # 計画作成日等
        ws1["A9"] = "居宅サービス計画作成（変更）日"
        ws1.merge_cells('B9:D9')
        ws1["B9"] = user_info['plan_date']
        ws1["E9"] = "初回居宅サービス計画作成日"
        ws1.merge_cells('F9:M9')
        ws1["F9"] = user_info['initial_plan_date']
        
        ws1["A10"] = "認定日"
        ws1.merge_cells('B10:D10')
        ws1["B10"] = user_info['certification_date']
        ws1["E10"] = "認定の有効期間"
        ws1.merge_cells('F10:M10')
        ws1["F10"] = f"{user_info['valid_from']} ～ {user_info['valid_to']}"
        
        # 要介護状態区分
        ws1["A12"] = "要介護状態区分"
        ws1.merge_cells('B12:M12')
        care_levels = ["要支援１", "要支援２", "要介護１", "要介護２", "要介護３", "要介護４", "要介護５"]
        care_level_checks = []
        for level in care_levels:
            if level == user_info['care_level']:
                care_level_checks.append(f"☑{level}")
            else:
                care_level_checks.append(f"□{level}")
        ws1["B12"] = " ・ ".join(care_level_checks)
        
        # 利用者及び家族の生活に対する意向
        ws1["A14"] = "利用者及び家族の生活に対する意向"
        ws1.merge_cells('B14:M19')
        ws1["B14"] = user_info['client_family_intentions']
        
        # 介護認定審査会の意見及びサービスの種類の指定
        ws1["A20"] = "介護認定審査会の意見及びサービスの種類の指定"
        ws1.merge_cells('B20:M23')
        ws1["B20"] = user_info['certification_opinion']
        
        # 総合的な援助の方針
        ws1["A24"] = "総合的な援助の方針"
        ws1.merge_cells('B24:M32')
        ws1["B24"] = user_info['support_policy']
        
        # 生活援助中心型の算定理由
        ws1["A33"] = "生活援助中心型の算定理由"
        ws1.merge_cells('B33:M33')
        ws1["B33"] = user_info['care_reason']
        
        # 同意欄
        ws1.merge_cells('A35:D35')
        ws1["A35"] = "居宅サービス計画について説明を受け、内容に同意し、交付を受けました。"
        ws1.merge_cells('E35:H35')
        ws1["E35"] = user_info['consent']['date']
        ws1.merge_cells('I35:M35')
        ws1["I35"] = f"氏名：{user_info['consent']['signature']}　印"
        
        # スタイル設定
        for row in ws1.rows:
            for cell in row:
                cell.font = Font(name='游ゴシック', size=10)
                cell.alignment = Alignment(wrap_text=True, vertical='center')
        
        # 列幅の設定
        ws1.column_dimensions['A'].width = 35
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws1.column_dimensions[col].width = 15
        
        # 行の高さ設定
        for i in range(1, 36):
            ws1.row_dimensions[i].height = 20
        
        # セルの罫線設定
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in ws1.rows:
            for cell in row:
                cell.border = thin_border
        
        # 第2表シート
        ws2 = wb.create_sheet("第2表")
        
        # タイトルと作成日
        ws2.merge_cells('A1:B1')
        ws2["A1"] = "第2表"
        ws2.merge_cells('C1:J1')
        ws2["C1"] = "居宅サービス計画書（２）"
        ws2.merge_cells('K1:M1')
        ws2["K1"] = f"作成年月日：{datetime.now().strftime('%Y年%m月%d日')}"
        
        # 利用者情報
        ws2["A3"] = "利用者名"
        ws2.merge_cells('B3:E3')
        ws2["B3"] = user_info['name']
        ws2["F3"] = "様"
        ws2.merge_cells('G3:I3')
        ws2["G3"] = "居宅サービス計画作成者"
        ws2.merge_cells('J3:M3')
        ws2["J3"] = user_info['care_manager']
        
        # テーブルヘッダー
        ws2.merge_cells('A5:A6')
        ws2["A5"] = "生活全般の解決すべき課題（ニーズ）"
        
        ws2.merge_cells('B5:E5')
        ws2["B5"] = "援助目標"
        ws2.merge_cells('B6:C6')
        ws2["B6"] = "長期目標（期間）"
        ws2.merge_cells('D6:E6')
        ws2["D6"] = "短期目標（期間）"
        
        ws2.merge_cells('F5:M5')
        ws2["F5"] = "援助内容"
        ws2["F6"] = "サービス内容"
        ws2["G6"] = "※1"
        ws2.merge_cells('H6:I6')
        ws2["H6"] = "サービス種別"
        ws2["J6"] = "※2"
        ws2["K6"] = "頻度"
        ws2.merge_cells('L6:M6')
        ws2["L6"] = "期間"
        
        # データ行の設定
        start_row = 7
        if 'care_plan_data' in st.session_state and st.session_state.care_plan_data.get('issues'):
            for i, issue in enumerate(st.session_state.care_plan_data['issues']):
                row = start_row + i
                
                # 課題
                ws2[f"A{row}"] = issue['needs']
                
                # 長期目標
                ws2.merge_cells(f'B{row}:C{row}')
                ws2[f"B{row}"] = f"{issue['long_term_goal']}\n（{issue['long_term_period']}）"
                
                # 短期目標
                ws2.merge_cells(f'D{row}:E{row}')
                ws2[f"D{row}"] = f"{issue['short_term_goal']}\n（{issue['short_term_period']}）"
                
                # サービス内容
                ws2[f"F{row}"] = issue['service_content']
                ws2[f"G{row}"] = issue['insurance_covered']
                
                # サービス種別
                ws2.merge_cells(f'H{row}:I{row}')
                ws2[f"H{row}"] = issue['service_type']
                
                ws2[f"J{row}"] = issue['service_provider']
                ws2[f"K{row}"] = issue['frequency']
                
                # 期間
                ws2.merge_cells(f'L{row}:M{row}')
                ws2[f"L{row}"] = issue['period']
        
        # フッター注釈
        ws2["A35"] = "※1 「保険給付対象かどうかの区分」について、保険給付対象内サービスについては○印を付す。"
        ws2["A36"] = "※2 「当該サービス提供を行う事業所」について記入する。"
        
        # スタイル設定
        # フォント設定
        for row in ws2.rows:
            for cell in row:
                cell.font = Font(name='游ゴシック', size=10)
                cell.alignment = Alignment(wrap_text=True, vertical='center')
        
        # ヘッダーセルのスタイル
        header_cells = ['A5', 'B5', 'F5', 'A6', 'B6', 'D6', 'F6', 'G6', 'H6', 'J6', 'K6', 'L6']
        for cell in header_cells:
            ws2[cell].font = Font(name='游ゴシック', size=10, bold=True)
            ws2[cell].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            ws2[cell].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        
        # 列幅の設定
        ws2.column_dimensions['A'].width = 30  # 課題列
        for col in ['B', 'C', 'D', 'E']:  # 目標列
            ws2.column_dimensions[col].width = 15
        ws2.column_dimensions['F'].width = 25  # サービス内容列
        ws2.column_dimensions['G'].width = 5   # ※1列
        for col in ['H', 'I']:  # サービス種別列
            ws2.column_dimensions[col].width = 12
        ws2.column_dimensions['J'].width = 5   # ※2列
        ws2.column_dimensions['K'].width = 10  # 頻度列
        for col in ['L', 'M']:  # 期間列
            ws2.column_dimensions[col].width = 8
        
        # 行の高さ設定
        ws2.row_dimensions[5].height = 30  # ヘッダー1行目
        ws2.row_dimensions[6].height = 30  # ヘッダー2行目
        for i in range(7, 35):  # データ行
            ws2.row_dimensions[i].height = 40
        
        # 罫線設定
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        
        # データ領域の罫線
        for row in range(5, 35):
            for col in range(1, 14):  # A から M まで
                cell = ws2.cell(row=row, column=col)
                if row in [5, 6]:  # ヘッダー行
                    cell.border = thick_border
                else:
                    cell.border = thin_border
        
        # 第3表シート
        ws3 = wb.create_sheet("第3表")
        
        # タイトルと作成日
        ws3.merge_cells('A1:B1')
        ws3["A1"] = "第3表"
        ws3.merge_cells('C1:J1')
        ws3["C1"] = "週間サービス計画表"
        ws3.merge_cells('K1:M1')
        ws3["K1"] = f"作成年月日：令和{int(datetime.now().year) - 2018}年{datetime.now().month}月{datetime.now().day}日"
        
        # 利用者情報
        ws3["A3"] = "利用者名"
        ws3.merge_cells('B3:E3')
        ws3["B3"] = user_info['name']
        ws3["F3"] = "様"
        
        # 時間帯区分（縦書き）
        ws3.merge_cells('A5:A12')
        ws3["A5"] = "午前"
        ws3.merge_cells('A13:A20')
        ws3["A13"] = "午後"
        ws3.merge_cells('A21:A24')
        ws3["A21"] = "深夜"
        
        # 曜日ヘッダー
        days = ["月", "火", "水", "木", "金", "土", "日"]
        for i, day in enumerate(days):
            col = chr(ord('C') + i)
            ws3[f"{col}4"] = day
        
        # 時間帯
        times = ["6:00", "8:00", "10:00", "12:00",  # 午前
                "14:00", "16:00", "18:00", "20:00", "22:00",  # 午後
                "0:00", "2:00", "4:00", "6:00"]  # 深夜
        
        for i, time in enumerate(times):
            row = i + 5
            ws3[f"B{row}"] = time
        
        # 主な日常生活上の活動
        ws3["K4"] = "主な日常生活上の活動"
        ws3.merge_cells('K4:M4')
        
        # 週単位以外のサービス
        ws3.merge_cells('A25:M25')
        ws3["A25"] = "週単位以外のサービス"
        ws3.merge_cells('A26:M28')
        
        # スタイル設定
        # フォント設定
        for row in ws3.rows:
            for cell in row:
                cell.font = Font(name='游ゴシック', size=10)
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        
        # 時間帯区分の縦書き設定
        for cell in [ws3["A5"], ws3["A13"], ws3["A21"]]:
            cell.alignment = Alignment(textRotation=255, horizontal='center', vertical='center')
        
        # 列幅の設定
        ws3.column_dimensions['A'].width = 5   # 時間帯区分
        ws3.column_dimensions['B'].width = 8   # 時間
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:  # 曜日列
            ws3.column_dimensions[col].width = 12
        for col in ['K', 'L', 'M']:  # 主な日常生活上の活動
            ws3.column_dimensions[col].width = 15
        
        # 行の高さ設定
        ws3.row_dimensions[1].height = 30  # タイトル行
        ws3.row_dimensions[3].height = 25  # 利用者情報行
        ws3.row_dimensions[4].height = 25  # 曜日ヘッダー行
        for i in range(5, 25):  # 時間割表
            ws3.row_dimensions[i].height = 30
        for i in range(25, 29):  # 週単位以外のサービス
            ws3.row_dimensions[i].height = 25
        
        # 罫線設定
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # データ領域の罫線
        for row in range(4, 29):
            for col in range(1, 14):  # A から M まで
                cell = ws3.cell(row=row, column=col)
                cell.border = thin_border
        
        # Excelファイルをバイトストリームとして保存
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer

    except Exception as e:
        st.error(f"エクセルファイル生成中にエラーが発生しました: {str(e)}")
        return None

def load_questions():
    """questions.ymlから質問設定を読み込む"""
    try:
        with open('data/questions.yml', 'r', encoding='utf-8') as file:
            return yaml.safe_load(file)
    except Exception as e:
        st.error(f"質問設定の読み込みに失敗しました: {str(e)}")
        return None

def generate_prompt(question, context):
    """質問テンプレートから実際のプロンプトを生成"""
    try:
        if 'prompt_template' in question:
            return question['prompt_template'].format(**context)
        elif 'prompt' in question:
            return question['prompt'].format(**context)
        else:
            st.error(f"質問 {question['id']} にプロンプトテンプレートが設定されていません")
            return None
    except KeyError as e:
        st.error(f"プロンプト生成に必要な値が不足しています: {str(e)}")
        return None
    except Exception as e:
        st.error(f"プロンプト生成中にエラーが発生しました: {str(e)}")
        return None

def get_response_with_retry(prompt, params):
    """OpenAI APIを使用してレスポンスを取得（リトライ機能付き）"""
    try:
        # パラメータの設定
        temperature = params.get("temperature", 0.7)
        max_tokens = params.get("max_tokens", 1000)
        
        # リクエストの送信
        response = client.chat.completions.create(
            model="gpt-4-turbo-preview",
            messages=[
                {"role": "system", "content": "あなたは経験豊富な介護支援専門員です。"},
                {"role": "user", "content": prompt}
            ],
            temperature=temperature,
            max_tokens=max_tokens
        )
        
        return response.choices[0].message.content
        
    except Exception as e:
        st.error(f"APIリクエスト中にエラーが発生しました: {str(e)}")
        return None

def generate_care_plan(user_info, adl_data, client_needs):
    """OpenAI APIを使用してケアプラン生成"""
    try:
        # 質問設定の読み込み
        questions = load_questions()
        if not questions:
            st.error("質問設定の読み込みに失敗しました")
            return None
        
        # 回答を格納する辞書
        all_results = {}
        
        # 利用者・家族の意向を生成
        st.write("利用者・家族の意向を生成中...")
        intentions = generate_intentions_text(user_info, adl_data)
        if not intentions:
            st.error("利用者・家族の意向の回答生成に失敗しました")
            return None
        all_results["intentions"] = intentions
        
        # 認定審査会の意見を生成
        st.write("認定審査会の意見を生成中...")
        certification_opinion = generate_certification_opinion(user_info, adl_data)
        if not certification_opinion:
            st.error("認定審査会の意見の回答生成に失敗しました")
            return None
        all_results["certification_opinion"] = certification_opinion
        
        # 総合的な援助の方針を生成
        st.write("総合的な援助の方針を生成中...")
        support_policy = generate_support_policy(user_info, adl_data, intentions, certification_opinion)
        if not support_policy:
            st.error("総合的な援助の方針の回答生成に失敗しました")
            return None
        all_results["support_policy"] = support_policy
        
        # 各質問に対して回答を生成（その他の質問があれば）
        for question in questions['questions']:
            if question['id'] not in all_results:  # すでに処理済みの質問はスキップ
                # titleキーがない場合はidを使用
                question_title = question.get('title', question['id'])
                st.write(f"処理中の質問: {question_title}")
                
                # プロンプトの生成
                with st.spinner(f"{question_title}のプロンプトを生成中..."):
                    # ADL項目の整形（care_goals用）
                    adl_items_text = ""
                    if question['id'] == 'care_goals':
                        adl_items_text = "\n".join([f"- {key}: {value}" for key, value in adl_data.items()])
                    
                    prompt = generate_prompt(question, {
                        "name": user_info['name'],
                        "care_level": user_info['care_level'],
                        "adl_data": pd.DataFrame([adl_data]).T.to_string(),
                        "client_needs": client_needs,
                        "intentions": intentions,
                        "certification_opinion": certification_opinion,
                        "adl_items": adl_items_text
                    })
                    
                    if not prompt:
                        st.error(f"{question_title}のプロンプト生成に失敗しました")
                        continue
                
                # APIリクエスト
                with st.spinner(f"{question_title}の回答を生成中..."):
                    params = {
                        "temperature": question.get("temperature", 0.7),
                        "max_tokens": question.get("max_tokens", 1000)
                    }
                    
                    response = get_response_with_retry(prompt, params)
                    if not response:
                        st.error(f"{question_title}の回答生成に失敗しました")
                        continue
                    
                    all_results[question['id']] = response
        
        # 結果の評価と整形
        evaluated_results = evaluate_response(all_results)
        formatted_results = format_care_plan(evaluated_results)
        
        return formatted_results
        
    except Exception as e:
        st.error(f"ケアプラン生成中にエラーが発生しました: {str(e)}")
        st.error(f"エラーの詳細:\n\n{traceback.format_exc()}")
        return None

def evaluate_response(response):
    """生成された回答の評価"""
    try:
        # ここでは単純に回答をそのまま返す
        return response
    except Exception as e:
        st.error(f"回答の評価中にエラーが発生しました: {str(e)}")
        return response

def format_care_plan(results):
    """ケアプランの整形"""
    try:
        # 整形されたケアプランを格納する辞書
        formatted_plan = {
            "intentions": results.get("intentions", ""),
            "certification_opinion": results.get("certification_opinion", ""),
            "support_policy": results.get("support_policy", "")
        }
        
        # その他の質問があれば追加
        for key, value in results.items():
            if key not in formatted_plan:
                formatted_plan[key] = value
        
        return formatted_plan
    except Exception as e:
        st.error(f"ケアプランの整形中にエラーが発生しました: {str(e)}")
        return results

def output_csv(results):
    """結果をCSVファイルに出力"""
    try:
        st.write("CSV出力を開始します")
        st.write("出力する結果:", results)
        
        # 現在の日時を取得
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # データフレームを作成
        data_list = []
        for qid, data in results.items():
            # データが文字列の場合（新しい形式）
            if isinstance(data, str):
                data_list.append({
                    'question_id': qid,
                    'title': qid,  # idをタイトルとして使用
                    'content': data
                })
            # データが辞書の場合（古い形式）
            elif isinstance(data, dict) and 'title' in data and 'content' in data:
                data_list.append({
                    'question_id': qid,
                    'title': data['title'],
                    'content': data['content']
                })
            # その他の形式
            else:
                data_list.append({
                    'question_id': qid,
                    'title': qid,
                    'content': str(data)
                })
        
        df = pd.DataFrame(data_list)
        
        st.write("作成されたデータフレーム:")
        st.write(df)
        
        # CSVファイルに保存
        output_path = f"output/care_plan_{timestamp}.csv"
        os.makedirs('output', exist_ok=True)
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        
        st.success(f"結果を {output_path} に保存しました")
        
        # 保存したファイルの内容を確認
        if os.path.exists(output_path):
            with open(output_path, 'r', encoding='utf-8-sig') as f:
                st.write("保存されたCSVファイルの内容:")
                st.write(f.read())
        
    except Exception as e:
        st.error(f"CSV出力中にエラーが発生しました: {str(e)}")
        st.write("エラーの詳細:", e)

def render_adl_input_section(items, category_name):
    """ADL入力セクションのレンダリング"""
    st.markdown(f"### {category_name}", unsafe_allow_html=True)
    category_data = {}
    
    for item in items:
        st.markdown("""
            <div style='
                border: 1px solid #e0e0e0;
                border-radius: 10px;
                padding: 15px;
                margin: 10px 0;
                background-color: #ffffff;
                box-shadow: 0 2px 4px rgba(0,0,0,0.05);
            '>
        """, unsafe_allow_html=True)
        
        st.markdown(f"#### {item}")
        status = st.selectbox(
            "状態を選択",
            ["要全介助", "一部介助", "見守り", "自立"],
            key=f"adl_{item}"
        )
        category_data[item] = status
        
        description = get_adl_description(item, status)
        if description:
            st.markdown(
                f"""<div style='
                    padding: 10px;
                    border-radius: 5px;
                    background-color: {get_adl_status_color(status)}15;
                    border-left: 5px solid {get_adl_status_color(status)};
                    margin-top: 10px;
                '>{description}</div>""",
                unsafe_allow_html=True
            )
        
        st.markdown("</div>", unsafe_allow_html=True)
    
    return category_data

def create_care_plan_pdf(user_info, adl_data, care_plan):
    """ケアプランをPDFファイルとして生成（レイアウト調整機能付き）"""
    try:
        # カスタムスタイルの定義
        styles = getSampleStyleSheet()
        
        # 日本語フォントの登録
        pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
        pdfmetrics.registerFont(UnicodeCIDFont('HeiseiMin-W3'))
        
        # カスタムスタイルの作成
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName='HeiseiKakuGo-W5',
            fontSize=16,
            alignment=1,  # 中央揃え
            spaceAfter=20
        )
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading2'],
            fontName='HeiseiKakuGo-W5',
            fontSize=12,
            alignment=0,  # 左揃え
            spaceAfter=10
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName='HeiseiMin-W3',
            fontSize=10,
            alignment=0,  # 左揃え
            leading=14
        )
        
        # PDFバッファの作成
        pdf_buffer = BytesIO()
        
        # PDFドキュメントの設定
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=A4,
            rightMargin=25*mm,
            leftMargin=25*mm,
            topMargin=25*mm,
            bottomMargin=25*mm
        )
        
        # ドキュメント要素の作成
        elements = []
        
        # タイトルと作成日
        elements.append(Paragraph("居宅サービス計画書（１）", title_style))
        elements.append(Paragraph(f"作成年月日：{datetime.now().strftime('%Y年%m月%d日')}", normal_style))
        elements.append(Spacer(1, 10*mm))
        
        # 認定区分
        certification_text = "計画区分："
        if user_info['plan_status']['initial']:
            certification_text += "☑"
        else:
            certification_text += "□"
        certification_text += "初回 "
        if user_info['plan_status']['introduced']:
            certification_text += "☑"
        else:
            certification_text += "□"
        certification_text += "紹介 "
        if user_info['plan_status']['continuous']:
            certification_text += "☑"
        else:
            certification_text += "□"
        certification_text += "継続"
        elements.append(Paragraph(certification_text, normal_style))
        
        # 認定状況
        status_text = "認定状況："
        if user_info['plan_status']['certified']:
            status_text += "☑"
        else:
            status_text += "□"
        status_text += "認定済 "
        if user_info['plan_status']['applying']:
            status_text += "☑"
        else:
            status_text += "□"
        status_text += "申請中"
        elements.append(Paragraph(status_text, normal_style))
        elements.append(Spacer(1, 5*mm))
        
        # 基本情報
        elements.append(Paragraph("■ 基本情報", header_style))
        elements.append(Paragraph(f"利用者名：{user_info['name']} 様", normal_style))
        elements.append(Paragraph(f"生年月日：{user_info['birth_date']}", normal_style))
        elements.append(Paragraph(f"住所：{user_info['address']}", normal_style))
        elements.append(Paragraph(f"居宅サービス計画作成者：{user_info['care_manager']}", normal_style))
        elements.append(Paragraph(f"事業所：{user_info['care_office']}", normal_style))
        elements.append(Paragraph(f"所在地：{user_info['office_address']}", normal_style))
        elements.append(Spacer(1, 5*mm))
        
        # 要介護状態区分
        elements.append(Paragraph("■ 要介護状態区分", header_style))
        elements.append(Paragraph(user_info['care_level'], normal_style))
        elements.append(Spacer(1, 5*mm))
        
        # 利用者・家族の意向
        elements.append(Paragraph("■ 利用者・家族の意向", header_style))
        elements.append(Paragraph(user_info['client_family_intentions'], normal_style))
        elements.append(Spacer(1, 5*mm))
        
        # 認定審査会意見
        elements.append(Paragraph("■ 認定審査会意見", header_style))
        elements.append(Paragraph(user_info['certification_opinion'], normal_style))
        elements.append(Spacer(1, 5*mm))
        
        # 総合的な援助の方針
        elements.append(Paragraph("■ 総合的な援助の方針", header_style))
        elements.append(Paragraph(user_info['support_policy'], normal_style))
        elements.append(Spacer(1, 5*mm))
        
        # 同意欄
        elements.append(Spacer(1, 10*mm))
        elements.append(Paragraph("居宅サービス計画について説明を受け、内容に同意し、交付を受けました。", normal_style))
        elements.append(Spacer(1, 5*mm))
        elements.append(Paragraph(f"日付：{user_info['consent']['date']}", normal_style))
        
        # 署名・印鑑の追加
        if 'signature_and_seal' in user_info:
            signature_data = user_info['signature_and_seal']['signature']
            seal_data = user_info['signature_and_seal']['seal']
            
            signature_text = "署名："
            if signature_data['type'] == "テキスト入力":
                signature_text += signature_data['data']
            elif signature_data['type'] == "画像アップロード" and signature_data['file']:
                # TODO: 署名画像の挿入処理
                pass
            
            elements.append(Paragraph(signature_text, normal_style))
        
        # PDFの生成
        doc.build(elements)
        pdf_buffer.seek(0)
        
        return pdf_buffer
        
    except Exception as e:
        st.error(f"PDF生成中にエラーが発生しました: {str(e)}")
        return None

def create_download_package(user_info, adl_data, care_plan, timestamp):
    """各形式のファイルを含むZIPパッケージを作成"""
    try:
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            # テキストファイル
            zf.writestr(
                f"care_plan_{timestamp}.txt",
                care_plan
            )
            
            # Excelファイル
            excel_buffer = create_care_plan_excel(user_info, adl_data, care_plan)
            if excel_buffer:
                zf.writestr(
                    f"care_plan_{timestamp}.xlsx",
                    excel_buffer.getvalue()
                )
            
            # PDFファイル
            pdf_buffer = create_care_plan_pdf(user_info, adl_data, care_plan)
            if pdf_buffer:
                zf.writestr(
                    f"care_plan_{timestamp}.pdf",
                    pdf_buffer.getvalue()
                )
        
        zip_buffer.seek(0)
        return zip_buffer
    
    except Exception as e:
        st.error(f"パッケージ作成中にエラーが発生しました: {str(e)}")
        return None

def preview_care_plan_table(user_info):
    """居宅サービス計画書（１）のプレビューを生成"""
    preview = f"""
    <div style='border: 1px solid #ddd; padding: 20px; border-radius: 5px; background-color: white;'>
        <h3 style='text-align: center; margin-bottom: 20px;'>居宅サービス計画書（１）</h3>
        <div style='text-align: right;'>作成年月日：{datetime.now().strftime('%Y年%m月%d日')}</div>
        
        <div style='margin-top: 20px;'>
            <div style='display: flex; justify-content: space-between;'>
                <div>
                    計画区分：
                    {' ☑' if user_info['plan_status']['initial'] else ' □'}初回
                    {' ☑' if user_info['plan_status']['introduced'] else ' □'}紹介
                    {' ☑' if user_info['plan_status']['continuous'] else ' □'}継続
                </div>
                <div>
                    認定状況：
                    {' ☑' if user_info['plan_status']['certified'] else ' □'}認定済
                    {' ☑' if user_info['plan_status']['applying'] else ' □'}申請中
                </div>
            </div>
        </div>
        
        <div style='margin-top: 20px;'>
            <table style='width: 100%; border-collapse: collapse;'>
                <tr>
                    <td style='width: 20%; padding: 8px; border: 1px solid #ddd;'>利用者名</td>
                    <td style='width: 50%; padding: 8px; border: 1px solid #ddd;'>{user_info['name']} 様</td>
                    <td style='width: 30%; padding: 8px; border: 1px solid #ddd;'>生年月日：{user_info['birth_date']}</td>
                </tr>
                <tr>
                    <td style='padding: 8px; border: 1px solid #ddd;'>住所</td>
                    <td colspan='2' style='padding: 8px; border: 1px solid #ddd;'>{user_info['address']}</td>
                </tr>
                <tr>
                    <td style='padding: 8px; border: 1px solid #ddd;'>居宅サービス計画作成者</td>
                    <td colspan='2' style='padding: 8px; border: 1px solid #ddd;'>{user_info['care_manager']}</td>
                </tr>
                <tr>
                    <td style='padding: 8px; border: 1px solid #ddd;'>事業所</td>
                    <td colspan='2' style='padding: 8px; border: 1px solid #ddd;'>{user_info['care_office']}<br>{user_info['office_address']}</td>
                </tr>
            </table>
        </div>
        
        <div style='margin-top: 20px;'>
            <h4>要介護状態区分</h4>
            <div style='padding: 10px; border: 1px solid #ddd;'>
                {user_info['care_level']}
            </div>
        </div>
        
        <div style='margin-top: 20px;'>
            <h4>利用者・家族の意向</h4>
            <div style='padding: 10px; border: 1px solid #ddd; min-height: 100px;'>
                {user_info['client_family_intentions']}
            </div>
        </div>
        
        <div style='margin-top: 20px;'>
            <h4>認定審査会意見</h4>
            <div style='padding: 10px; border: 1px solid #ddd; min-height: 100px;'>
                {user_info['certification_opinion']}
            </div>
        </div>
        
        <div style='margin-top: 20px;'>
            <h4>総合的な援助の方針</h4>
            <div style='padding: 10px; border: 1px solid #ddd; min-height: 100px;'>
                {user_info['support_policy']}
            </div>
        </div>
        
        <div style='margin-top: 20px;'>
            <div style='text-align: center;'>
                <p>居宅サービス計画について説明を受け、内容に同意し、交付を受けました。</p>
                <div style='margin-top: 10px;'>
                    {user_info['consent']['date']}<br>
                    氏名：{user_info['consent']['signature']} 印
                </div>
            </div>
        </div>
    </div>
    """
    return preview

def upload_signature_and_seal():
    """電子署名と印鑑画像のアップロード機能"""
    st.subheader("電子署名・印鑑設定")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### 電子署名")
        signature_type = st.radio(
            "署名方法を選択",
            ["手書き署名", "テキスト入力", "画像アップロード"]
        )
        
        if signature_type == "手書き署名":
            st.markdown("手書き署名エリア（実装予定）")
            # TODO: Canvas要素を使用した手書き署名機能の実装
        elif signature_type == "テキスト入力":
            signature_text = st.text_input("署名を入力")
            font_family = st.selectbox(
                "フォントを選択",
                ["明朝体", "ゴシック体", "行書体"]
            )
        else:
            signature_file = st.file_uploader("署名画像をアップロード", type=["png", "jpg", "jpeg"])
    
    with col2:
        st.markdown("### 印鑑画像")
        seal_type = st.radio(
            "印鑑タイプを選択",
            ["印鑑画像をアップロード", "デジタル印鑑を作成"]
        )
        
        if seal_type == "印鑑画像をアップロード":
            seal_file = st.file_uploader("印鑑画像をアップロード", type=["png", "jpg", "jpeg"])
        else:
            seal_text = st.text_input("印鑑に入れる文字")
            seal_color = st.color_picker("印鑑の色を選択", "#FF0000")
            seal_style = st.selectbox(
                "印鑑スタイルを選択",
                ["丸印", "角印", "楕円印"]
            )
    
    return {
        "signature": {
            "type": signature_type,
            "data": signature_text if signature_type == "テキスト入力" else None,
            "font": font_family if signature_type == "テキスト入力" else None,
            "file": signature_file if signature_type == "画像アップロード" else None
        },
        "seal": {
            "type": seal_type,
            "text": seal_text if seal_type == "デジタル印鑑を作成" else None,
            "color": seal_color if seal_type == "デジタル印鑑を作成" else None,
            "style": seal_style if seal_type == "デジタル印鑑を作成" else None,
            "file": seal_file if seal_type == "印鑑画像をアップロード" else None
        }
    }

def preview_care_plan_table_2(care_plan_data):
    """居宅サービス計画書（2）のプレビューを生成"""
    preview = f"""
    <div style='border: 1px solid #ddd; padding: 20px; border-radius: 5px; background-color: white;'>
        <div style='display: flex; justify-content: space-between; align-items: center;'>
            <div style='width: 10%;'>第2表</div>
            <div style='width: 70%; text-align: center;'>居宅サービス計画書（2）</div>
            <div style='width: 20%; text-align: right;'>作成年月日：{datetime.now().strftime('%Y年%m月%d日')}</div>
        </div>
        
        <div style='margin-top: 20px; display: flex; justify-content: space-between;'>
            <div>利用者名：{care_plan_data['user_info']['name']} 様</div>
            <div>居宅サービス計画作成者：{care_plan_data['user_info']['care_manager']}</div>
        </div>
        
        <table style='width: 100%; border-collapse: collapse; margin-top: 20px;'>
            <tr>
                <th style='border: 1px solid #ddd; padding: 8px; width: 20%; background-color: #f8f9fa;' rowspan='2'>
                    生活全般の解決すべき課題（ニーズ）
                </th>
                <th style='border: 1px solid #ddd; padding: 8px; background-color: #f8f9fa;' colspan='4'>
                    援助目標
                </th>
                <th style='border: 1px solid #ddd; padding: 8px; background-color: #f8f9fa;' colspan='6'>
                    援助内容
                </th>
            </tr>
            <tr>
                <th style='border: 1px solid #ddd; padding: 8px; background-color: #f8f9fa;' colspan='2'>長期目標（期間）</th>
                <th style='border: 1px solid #ddd; padding: 8px; background-color: #f8f9fa;' colspan='2'>短期目標（期間）</th>
                <th style='border: 1px solid #ddd; padding: 8px; background-color: #f8f9fa;'>サービス内容</th>
                <th style='border: 1px solid #ddd; padding: 8px; background-color: #f8f9fa;'>※1</th>
                <th style='border: 1px solid #ddd; padding: 8px; background-color: #f8f9fa;'>サービス種別</th>
                <th style='border: 1px solid #ddd; padding: 8px; background-color: #f8f9fa;'>※2</th>
                <th style='border: 1px solid #ddd; padding: 8px; background-color: #f8f9fa;'>頻度</th>
                <th style='border: 1px solid #ddd; padding: 8px; background-color: #f8f9fa;'>期間</th>
            </tr>
    """
    
    # 課題ごとのデータを表示
    for issue in care_plan_data.get('issues', []):
        preview += f"""
            <tr>
                <td style='border: 1px solid #ddd; padding: 8px;'>{issue['needs']}</td>
                <td style='border: 1px solid #ddd; padding: 8px;' colspan='2'>{issue['long_term_goal']}<br>（{issue['long_term_period']}）</td>
                <td style='border: 1px solid #ddd; padding: 8px;' colspan='2'>{issue['short_term_goal']}<br>（{issue['short_term_period']}）</td>
                <td style='border: 1px solid #ddd; padding: 8px;'>{issue['service_content']}</td>
                <td style='border: 1px solid #ddd; padding: 8px; text-align: center;'>{issue['insurance_covered']}</td>
                <td style='border: 1px solid #ddd; padding: 8px;'>{issue['service_type']}</td>
                <td style='border: 1px solid #ddd; padding: 8px;'>{issue['service_provider']}</td>
                <td style='border: 1px solid #ddd; padding: 8px;'>{issue['frequency']}</td>
                <td style='border: 1px solid #ddd; padding: 8px;'>{issue['period']}</td>
            </tr>
        """
    
    preview += """
        </table>
        
        <div style='margin-top: 20px; font-size: 0.9em;'>
            <p>※1 「保険給付対象かどうかの区分」について、保険給付対象内サービスについては○印を付す。</p>
            <p>※2 「当該サービス提供を行う事業所」について記入する。</p>
        </div>
    </div>
    """
    return preview

# 課題テンプレートの定義
ISSUE_TEMPLATES = {
    "ADL関連": [
        {
            "needs": "食事摂取の自立支援",
            "long_term_goal": "安全に自力で食事が摂取できる",
            "long_term_period": "6ヶ月",
            "short_term_goal": "食事動作の改善と誤嚥予防",
            "short_term_period": "3ヶ月",
            "service_content": "食事介助、姿勢調整、食形態の工夫",
            "service_type": "訪問介護",
            "frequency": "毎日3回",
            "period": "3ヶ月"
        },
        {
            "needs": "入浴の安全確保",
            "long_term_goal": "安全な入浴動作の確立",
            "long_term_period": "6ヶ月",
            "short_term_goal": "介助を受けながら安全に入浴できる",
            "short_term_period": "3ヶ月",
            "service_content": "入浴介助、浴室環境整備",
            "service_type": "訪問入浴介護",
            "frequency": "週2回",
            "period": "3ヶ月"
        }
    ],
    "IADL関連": [
        {
            "needs": "服薬管理の支援",
            "long_term_goal": "確実な服薬管理の実現",
            "long_term_period": "6ヶ月",
            "short_term_goal": "薬の仕分けと服用時間の理解",
            "short_term_period": "3ヶ月",
            "service_content": "服薬管理指導、薬の仕分け支援",
            "service_type": "居宅療養管理指導",
            "frequency": "週1回",
            "period": "3ヶ月"
        }
    ],
    "社会参加": [
        {
            "needs": "外出機会の確保",
            "long_term_goal": "地域活動への参加再開",
            "long_term_period": "6ヶ月",
            "short_term_goal": "デイサービスでの活動参加",
            "short_term_period": "3ヶ月",
            "service_content": "通所介護サービスの利用",
            "service_type": "通所介護",
            "frequency": "週3回",
            "period": "3ヶ月"
        }
    ]
}

# サービス提供事業所のデータ
SERVICE_PROVIDERS = {
    "訪問介護": [
        "ヘルパーステーションEGAO",
        "訪問介護ステーションすまいる",
        "ホームヘルプサービスあい"
    ],
    "訪問入浴介護": [
        "訪問入浴サービスEGAO",
        "モバイルバスケアセンター",
        "訪問入浴ステーションゆとり"
    ],
    "訪問看護": [
        "訪問看護ステーションEGAO",
        "ナースステーションほほえみ",
        "訪問看護ステーションはあと"
    ],
    "通所介護": [
        "デイサービスEGAO",
        "デイサービスセンターわかば",
        "リハビリデイサービスすこやか"
    ]
}

# 頻度パターンの定義
FREQUENCY_PATTERNS = {
    "訪問介護": [
        "毎日1回",
        "毎日2回",
        "毎日3回",
        "週3回",
        "週2回",
        "週1回"
    ],
    "訪問入浴介護": [
        "週2回",
        "週1回"
    ],
    "通所介護": [
        "週3回",
        "週2回",
        "週1回"
    ],
    "訪問看護": [
        "週2回",
        "週1回",
        "2週間に1回"
    ]
}

def edit_care_plan_issue():
    """課題の編集機能"""
    st.subheader("課題の編集")
    
    # サービス種別のリスト
    service_types = [
        "訪問介護", "訪問入浴介護", "訪問看護", "訪問リハビリテーション",
        "居宅療養管理指導", "通所介護", "通所リハビリテーション",
        "短期入所生活介護", "短期入所療養介護", "福祉用具貸与",
        "特定福祉用具販売", "住宅改修", "居宅介護支援"
    ]
    
    # 保険給付対象サービスのリスト
    insurance_covered_services = [
        "訪問介護", "訪問入浴介護", "訪問看護", "訪問リハビリテーション",
        "居宅療養管理指導", "通所介護", "通所リハビリテーション",
        "短期入所生活介護", "短期入所療養介護", "福祉用具貸与"
    ]
    
    # テンプレートからの課題追加
    st.markdown("### テンプレートから追加")
    col1, col2 = st.columns(2)
    with col1:
        template_category = st.selectbox(
            "カテゴリー選択",
            list(ISSUE_TEMPLATES.keys())
        )
    with col2:
        if template_category:
            template_issues = ISSUE_TEMPLATES[template_category]
            template_names = [issue["needs"] for issue in template_issues]
            selected_template = st.selectbox(
                "テンプレート選択",
                template_names
            )
            if st.button("テンプレートを追加"):
                if 'care_plan_data' not in st.session_state:
                    st.session_state.care_plan_data = {'issues': []}
                
                template_issue = next(
                    issue for issue in template_issues 
                    if issue["needs"] == selected_template
                )
                st.session_state.care_plan_data['issues'].append(template_issue.copy())
                st.success("テンプレートを追加しました")
                st.rerun()
    
    # 新規課題の追加
    st.markdown("### 新規課題の追加")
    if st.button("新規課題を追加"):
        if 'care_plan_data' not in st.session_state:
            st.session_state.care_plan_data = {'issues': []}
        
        new_issue = {
            'needs': '',
            'long_term_goal': '',
            'long_term_period': '6ヶ月',
            'short_term_goal': '',
            'short_term_period': '3ヶ月',
            'service_content': '',
            'service_type': service_types[0],
            'service_provider': '',
            'frequency': '',
            'period': '',
            'insurance_covered': ''
        }
        st.session_state.care_plan_data['issues'].append(new_issue)
        st.rerun()
    
    # 課題の並び替え
    if 'care_plan_data' in st.session_state and len(st.session_state.care_plan_data['issues']) > 1:
        st.markdown("### 課題の並び替え")
        col1, col2 = st.columns(2)
        with col1:
            issue_to_move = st.selectbox(
                "移動する課題を選択",
                [f"課題{i+1}: {issue['needs']}" for i, issue in enumerate(st.session_state.care_plan_data['issues'])]
            )
        with col2:
            new_position = st.number_input(
                "新しい位置",
                min_value=1,
                max_value=len(st.session_state.care_plan_data['issues']),
                value=1
            )
        
        if st.button("並び替えを実行"):
            current_index = int(issue_to_move.split(':')[0].replace('課題', '')) - 1
            new_index = new_position - 1
            issues = st.session_state.care_plan_data['issues']
            issue = issues.pop(current_index)
            issues.insert(new_index, issue)
            st.success("課題の順序を変更しました")
            st.rerun()
    
    # 既存の課題を編集
    if 'care_plan_data' in st.session_state and st.session_state.care_plan_data['issues']:
        st.markdown("### 課題の編集")
        for i, issue in enumerate(st.session_state.care_plan_data['issues']):
            with st.expander(f"課題 {i+1}: {issue['needs'] or '(未入力)'}"):
                col1, col2 = st.columns(2)
                
                with col1:
                    issue['needs'] = st.text_area(
                        "生活全般の解決すべき課題（ニーズ）",
                        issue['needs'],
                        key=f"needs_{i}"
                    )
                    issue['long_term_goal'] = st.text_area(
                        "長期目標",
                        issue['long_term_goal'],
                        key=f"long_term_goal_{i}"
                    )
                    issue['long_term_period'] = st.text_input(
                        "長期目標の期間",
                        issue['long_term_period'],
                        key=f"long_term_period_{i}"
                    )
                
                with col2:
                    issue['short_term_goal'] = st.text_area(
                        "短期目標",
                        issue['short_term_goal'],
                        key=f"short_term_goal_{i}"
                    )
                    issue['short_term_period'] = st.text_input(
                        "短期目標の期間",
                        issue['short_term_period'],
                        key=f"short_term_period_{i}"
                    )
                
                col3, col4 = st.columns(2)
                
                with col3:
                    issue['service_content'] = st.text_area(
                        "サービス内容",
                        issue['service_content'],
                        key=f"service_content_{i}"
                    )
                    issue['service_type'] = st.selectbox(
                        "サービス種別",
                        service_types,
                        key=f"service_type_{i}"
                    )
                    # 保険給付対象の自動チェック
                    issue['insurance_covered'] = "○" if issue['service_type'] in insurance_covered_services else ""
                    
                    # サービス提供事業所の候補表示
                    if issue['service_type'] in SERVICE_PROVIDERS:
                        provider_options = [""] + SERVICE_PROVIDERS[issue['service_type']]
                        issue['service_provider'] = st.selectbox(
                            "サービス提供事業所",
                            provider_options,
                            key=f"provider_{i}"
                        )
                    else:
                        issue['service_provider'] = st.text_input(
                            "サービス提供事業所",
                            issue['service_provider'],
                            key=f"provider_{i}"
                        )
                
                with col4:
                    # 頻度パターンの選択
                    if issue['service_type'] in FREQUENCY_PATTERNS:
                        frequency_options = [""] + FREQUENCY_PATTERNS[issue['service_type']]
                        selected_frequency = st.selectbox(
                            "頻度パターン",
                            frequency_options,
                            key=f"frequency_pattern_{i}"
                        )
                        if selected_frequency:
                            issue['frequency'] = selected_frequency
                    else:
                        issue['frequency'] = st.text_input(
                            "頻度",
                            issue['frequency'],
                            key=f"frequency_{i}"
                        )
                    
                    issue['period'] = st.text_input(
                        "期間",
                        issue['period'],
                        key=f"period_{i}"
                    )
                
                if st.button("この課題を削除", key=f"delete_{i}"):
                    st.session_state.care_plan_data['issues'].pop(i)
                    st.rerun()

def validate_client_info(user_info, adl_data, client_needs):
    """クライアント情報の検証"""
    if not user_info or not adl_data or not client_needs:
        st.error("クライアントの情報が不足しています。")
        return False
    return True

def define_rules_and_goals(adl_data):
    """ルールと目標の定義"""
    stages = calculate_stages(adl_data)
    motivation = calculate_motivation(adl_data)
    return {
        'stages': stages,
        'motivation': motivation,
        'goals': generate_goals(stages, motivation)
    }

def evaluate_care_plan(care_plan):
    """ケアプランの品質評価"""
    score = 100
    required_elements = [
        "第1表", "第2表", "第3表",
        "睡眠改善課題", "栄養改善課題",
        "長期目標", "短期目標"
    ]
    
    for element in required_elements:
        if element not in care_plan:
            score -= 10
    
    return score

def format_final_care_plan(care_plan):
    """最終的なケアプランのフォーマット"""
    # ケアプランの形式を整える
    formatted_plan = care_plan.replace("\n\n", "\n")
    return formatted_plan

def update_knowledge_base(care_plan, quality_score):
    """知識ベースの更新"""
    # 将来の改善のために情報を保存
    if 'care_plan_quality_history' not in st.session_state:
        st.session_state.care_plan_quality_history = []
    
    st.session_state.care_plan_quality_history.append({
        'timestamp': datetime.now(),
        'quality_score': quality_score,
        'improvements_needed': quality_score < 100
    })

def calculate_stages(adl_data):
    """ADLデータからステージを計算"""
    stage_weights = {
        "要全介助": 1,
        "一部介助": 2,
        "見守り": 3,
        "自立": 4
    }
    
    # 各ADL項目のステージ値を計算
    stage_values = {}
    for item, status in adl_data.items():
        stage_values[item] = stage_weights.get(status, 1)
    
    # 重要度に基づいて総合ステージを計算
    priority_items = {
        "移動": 1.5,
        "排泄": 1.3,
        "食事": 1.2,
        "入浴": 1.1
    }
    
    total_weight = sum(priority_items.values()) + (len(adl_data) - len(priority_items))
    weighted_sum = 0
    
    for item, stage in stage_values.items():
        weight = priority_items.get(item, 1.0)
        weighted_sum += stage * weight
    
    average_stage = weighted_sum / total_weight
    
    # ステージを1-5の範囲に変換
    final_stage = int((average_stage / 4) * 5)
    return max(1, min(5, final_stage))

def calculate_motivation(adl_data):
    """ADLデータからモチベーションレベルを計算"""
    motivation_indicators = {
        "コミュニケーション": 1.5,
        "認知機能": 1.3,
        "睡眠": 1.2
    }
    
    status_scores = {
        "要全介助": 0,
        "一部介助": 1,
        "見守り": 1.5,
        "自立": 2
    }
    
    total_weight = sum(motivation_indicators.values())
    weighted_sum = 0
    
    for item, weight in motivation_indicators.items():
        if item in adl_data:
            status = adl_data[item]
            score = status_scores.get(status, 0)
            weighted_sum += score * weight
    
    motivation_level = weighted_sum / total_weight
    return round(motivation_level, 1)

def generate_goals(stages, motivation):
    """ステージとモチベーションに基づいて目標を生成"""
    goals = {
        "short_term": [],
        "long_term": []
    }
    
    # ステージに基づく目標設定
    stage_goals = {
        1: {
            "short": "基本的なADLの安定",
            "long": "一部介助レベルへの改善"
        },
        2: {
            "short": "介助量の軽減",
            "long": "見守りレベルへの改善"
        },
        3: {
            "short": "見守り場面の特定",
            "long": "部分的な自立達成"
        },
        4: {
            "short": "自立範囲の拡大",
            "long": "完全自立の維持"
        },
        5: {
            "short": "現状機能の維持",
            "long": "社会参加の促進"
        }
    }
    
    # モチベーションに基づく追加目標
    motivation_goals = {
        0: {
            "short": "基本的なコミュニケーションの確立",
            "long": "意思表示の改善"
        },
        1: {
            "short": "日常的な意思疎通の向上",
            "long": "積極的な活動参加"
        },
        2: {
            "short": "社会的交流の促進",
            "long": "自己実現の支援"
        }
    }
    
    # 目標の設定
    stage_level = min(5, max(1, stages))
    motivation_level = min(2, max(0, int(motivation)))
    
    goals["short_term"].extend([
        stage_goals[stage_level]["short"],
        motivation_goals[motivation_level]["short"]
    ])
    
    goals["long_term"].extend([
        stage_goals[stage_level]["long"],
        motivation_goals[motivation_level]["long"]
    ])
    
    return goals

def generate_intentions_text(user_info, adl_data):
    """利用者・家族の意向のテキストを生成"""
    try:
        prompt = f"""
あなたは経験豊富な介護支援専門員です。以下の情報を元に、利用者・家族の生活に対する意向を記載してください。

利用者情報：
- 氏名：{user_info['name']}様
- 要介護度：{user_info['care_level']}
- ADL状態：
{pd.DataFrame([adl_data]).T.to_string()}

以下の点に注意して記載してください：
1. 利用者本人の意向を最優先に記載
2. 家族の意向も考慮
3. 現実的で具体的な内容
4. 本人の強みや残存機能を活かした内容
5. 社会参加や生きがいに関する内容も含める

出力形式：
【利用者本人の意向】
・
・
・

【家族の意向】
・
・
・
"""
        # OpenAIクライアントを使用
        response = client.chat.completions.create(
            model="gpt-4-turbo-preview",
            messages=[
                {"role": "system", "content": "あなたは経験豊富な介護支援専門員です。"},
                {"role": "user", "content": prompt}
            ],
            temperature=0.7,
            max_tokens=1000
        )
        
        return response.choices[0].message.content
        
    except Exception as e:
        st.error(f"テキスト生成中にエラーが発生しました: {str(e)}")
        return None

def generate_certification_opinion(user_info, adl_data):
    """認定審査会の意見を生成"""
    try:
        prompt = f"""
あなたは介護認定審査会の委員です。以下の情報を元に、介護認定審査会の意見を記載してください。

利用者情報：
- 氏名：{user_info['name']}様
- 要介護度：{user_info['care_level']}
- ADL状態：
{pd.DataFrame([adl_data]).T.to_string()}

以下の点に注意して記載してください：
1. 要介護状態の原因疾患や障害の状況
2. 介護の手間や必要な支援の内容
3. 改善可能性や予防の視点
4. 医学的管理の必要性
5. 推奨されるサービスの種類

出力形式：
1. 要介護状態の状況：
・

2. 改善可能性と予防：
・

3. 医学的管理の必要性：
・

4. 推奨サービス：
・
"""
        # OpenAIクライアントを使用
        response = client.chat.completions.create(
            model="gpt-4-turbo-preview",
            messages=[
                {"role": "system", "content": "あなたは介護認定審査会の委員です。"},
                {"role": "user", "content": prompt}
            ],
            temperature=0.7,
            max_tokens=1000
        )
        
        return response.choices[0].message.content
        
    except Exception as e:
        st.error(f"テキスト生成中にエラーが発生しました: {str(e)}")
        return None

def generate_support_policy(user_info, adl_data, intentions, certification_opinion):
    """総合的な援助の方針を生成"""
    try:
        prompt = f"""
あなたは経験豊富な介護支援専門員です。以下の情報を元に、総合的な援助の方針を記載してください。

利用者情報：
- 氏名：{user_info['name']}様
- 要介護度：{user_info['care_level']}
- ADL状態：
{pd.DataFrame([adl_data]).T.to_string()}

利用者・家族の意向：
{intentions}

認定審査会の意見：
{certification_opinion}

以下の点に注意して記載してください：
1. ICFの視点（心身機能・身体構造、活動、参加）
2. 短期・長期の目標を明確に
3. 具体的なサービス内容と期待される効果
4. リスク管理と予防的視点
5. 多職種連携の方針
6. モニタリング方法

出力形式：
【総合的な援助の方針】
・
・
・
"""
        # OpenAIクライアントを使用
        response = client.chat.completions.create(
            model="gpt-4-turbo-preview",
            messages=[
                {"role": "system", "content": "あなたは経験豊富な介護支援専門員です。"},
                {"role": "user", "content": prompt}
            ],
            temperature=0.7,
            max_tokens=1500
        )
        
        return response.choices[0].message.content
        
    except Exception as e:
        st.error(f"テキスト生成中にエラーが発生しました: {str(e)}")
        return None

def export_care_plan_to_excel_template(user_info, adl_data, care_plan):
    """
    care_plan_1.xlsテンプレートを使用してケアプランをExcelに出力
    """
    try:
        # テンプレートファイルのパス
        template_path = "document/care_plan_1.xls"
        
        # outputディレクトリが存在しない場合は作成
        os.makedirs("output", exist_ok=True)
        
        # 現在の日時を取得してファイル名に使用
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"output/care_plan_{timestamp}.xlsx"  # .xlsxに変更
        
        # pandasを使用してテンプレートを読み込む
        df = pd.read_excel(template_path, engine='xlrd')
        
        # 新しいExcelファイルを作成
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # データフレームを書き込む
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # ワークシートを取得
            workbook = writer.book
            worksheet = workbook['Sheet1']
            
            # 基本情報の入力
            worksheet['C3'] = user_info.get('name', '')  # 利用者名
            worksheet['C4'] = user_info.get('birth_date', '')  # 生年月日
            worksheet['C5'] = user_info.get('address', '')  # 住所
            worksheet['C6'] = user_info.get('care_level', '')  # 要介護度
            
            # ケアプラン情報の入力
            worksheet['C10'] = care_plan.get('intentions', '')  # 利用者・家族の意向
            worksheet['C12'] = care_plan.get('certification_opinion', '')  # 認定審査会の意見
            worksheet['C14'] = care_plan.get('support_policy', '')  # 総合的な援助の方針
            
            # care_goalsがある場合は入力
            if 'care_goals' in care_plan:
                worksheet['C16'] = care_plan.get('care_goals', '')  # ケア目標
            
            # service_planがある場合は入力
            if 'service_plan' in care_plan:
                worksheet['C18'] = care_plan.get('service_plan', '')  # サービス計画
        
        st.success(f"ケアプランをExcelに出力しました: {output_path}")
        return output_path
    except Exception as e:
        st.error(f"Excelへの出力中にエラーが発生しました: {str(e)}")
        st.error(f"エラーの詳細:\n\n{traceback.format_exc()}")
        return None

def main():
    st.markdown("""
        <h1 style='color: #1E88E5; font-size: 32px;'>
            EGAO-AI ケアプラン作成支援システム
        </h1>
    """, unsafe_allow_html=True)
    
    # サイドバーのスタイリング
    with st.sidebar:
        st.markdown("""
            <h2 style='color: #333; font-size: 24px; font-weight: 600;'>
                メニュー
            </h2>
        """, unsafe_allow_html=True)
        page = st.radio(
            "選択してください",
            ["基本情報入力", "ADLデータ入力", "ケアプラン生成", "履歴管理"]
        )

    # ADLカテゴリーの定義を更新
    adl_categories = {
        "🚶 基本動作": ["食事", "排泄", "入浴", "移動", "着替え", "整容"],
        "🧠 認知・コミュニケーション": ["コミュニケーション", "認知機能", "睡眠"],
        "🏠 社会生活": ["服薬管理", "金銭管理", "買い物"]
    }

    # 各ページのヘッダースタイリング
    if page == "基本情報入力":
        st.markdown("""
            <h2 style='color: #333; font-size: 28px; font-weight: 600; margin-bottom: 20px;'>
                居宅サービス計画書（１）
            </h2>
        """, unsafe_allow_html=True)
        
        # 認定区分
        col1, col2, col3 = st.columns(3)
        with col1:
            st.subheader("計画区分")
            initial_plan = st.checkbox("初回")
            introduced_plan = st.checkbox("紹介")
            continuous_plan = st.checkbox("継続")
        with col2:
            st.subheader("認定状況")
            certified = st.checkbox("認定済")
            applying = st.checkbox("申請中")
        with col3:
            st.markdown(f"""
                <div style='background-color: #f0f2f6; padding: 10px; border-radius: 5px;'>
                    <p style='margin: 0;'>作成年月日：{datetime.now().strftime('%Y年%m月%d日')}</p>
                </div>
            """, unsafe_allow_html=True)
        
        st.divider()
        
        # 基本情報入力
        st.subheader("基本情報")
        col1, col2 = st.columns(2)
        with col1:
            name = st.text_input("利用者名")
            min_date = datetime(1900, 1, 1)
            birth_date = st.date_input(
                "生年月日",
                min_value=min_date,
                max_value=datetime.now()
            )
            address = st.text_input("住所")
            care_manager = st.text_input("居宅サービス計画作成者氏名")
            care_office = st.text_input("居宅介護支援事業者・事業所名")
            office_address = st.text_input("事業所所在地")
            
        with col2:
            plan_date = st.date_input("居宅サービス計画作成（変更）日")
            initial_plan_date = st.date_input("初回居宅サービス計画作成日")
            certification_date = st.date_input("認定日")
            valid_from = st.date_input("認定の有効期間（開始）")
            valid_to = st.date_input("認定の有効期間（終了）")
        
        st.divider()
        
        # 要介護状態区分
        st.subheader("要介護状態区分")
        care_level = st.radio(
            "要介護度",
            ["要支援１", "要支援２", "要介護１", "要介護２", "要介護３", "要介護４", "要介護５"],
            horizontal=True
        )
        
        st.divider()
        
        # 利用者及び家族の生活に対する意向
        st.subheader("利用者及び家族の生活に対する意向")
        col1, col2 = st.columns([3, 1])
        with col1:
            if 'client_family_intentions' not in st.session_state:
                st.session_state.client_family_intentions = ""
            
            client_family_intentions = st.text_area(
                "利用者・家族の意向を入力してください",
                value=st.session_state.client_family_intentions,
                height=150,
                key="intentions_textarea"
            )
        with col2:
            if st.button("AIで生成", key="generate_intentions"):
                if 'user_info' in st.session_state and 'adl_data' in st.session_state:
                    with st.spinner("生成中..."):
                        generated_text = generate_intentions_text(
                            st.session_state.user_info,
                            st.session_state.adl_data
                        )
                        if generated_text:
                            st.session_state.client_family_intentions = generated_text
                            st.rerun()
                else:
                    st.warning("基本情報とADLデータを先に入力してください")
        
        # 介護認定審査会の意見及びサービスの種類の指定
        st.subheader("介護認定審査会の意見及びサービスの種類の指定")
        col1, col2 = st.columns([3, 1])
        with col1:
            if 'certification_opinion' not in st.session_state:
                st.session_state.certification_opinion = ""
            
            certification_opinion = st.text_area(
                "介護認定審査会の意見を入力してください",
                value=st.session_state.certification_opinion,
                height=100,
                key="opinion_textarea"
            )
        with col2:
            if st.button("AIで生成", key="generate_opinion"):
                if 'user_info' in st.session_state and 'adl_data' in st.session_state:
                    with st.spinner("生成中..."):
                        generated_text = generate_certification_opinion(
                            st.session_state.user_info,
                            st.session_state.adl_data
                        )
                        if generated_text:
                            st.session_state.certification_opinion = generated_text
                            st.rerun()
                else:
                    st.warning("基本情報とADLデータを先に入力してください")
        
        # 総合的な援助の方針
        st.subheader("総合的な援助の方針")
        col1, col2 = st.columns([3, 1])
        with col1:
            if 'support_policy' not in st.session_state:
                st.session_state.support_policy = ""
            
            support_policy = st.text_area(
                "総合的な援助の方針を入力してください",
                value=st.session_state.support_policy,
                height=150,
                key="policy_textarea"
            )
        with col2:
            if st.button("AIで生成", key="generate_policy"):
                if 'user_info' in st.session_state and 'adl_data' in st.session_state:
                    with st.spinner("生成中..."):
                        generated_text = generate_support_policy(
                            st.session_state.user_info,
                            st.session_state.adl_data,
                            st.session_state.client_family_intentions,
                            st.session_state.certification_opinion
                        )
                        if generated_text:
                            st.session_state.support_policy = generated_text
                            st.rerun()
                else:
                    st.warning("基本情報とADLデータを先に入力してください")
        
        st.divider()
        
        # 生活援助中心型の算定理由
        st.subheader("生活援助中心型の算定理由")
        care_reason = st.radio(
            "算定理由",
            ["1. 一人暮らし", "2. 家族等が障害、疾病等", "3. その他"],
            horizontal=True
        )
        if care_reason == "3. その他":
            other_reason = st.text_input("その他の理由を入力してください")
        
        st.divider()
        
        # 同意欄
        st.subheader("同意欄")
        st.markdown("""
            <div style='background-color: #f0f2f6; padding: 15px; border-radius: 5px;'>
                <p>居宅サービス計画について説明を受け、内容に同意し、交付を受けました。</p>
            </div>
        """, unsafe_allow_html=True)
        
        consent_date = st.date_input("同意日")
        signature = st.text_input("署名")
        
        # プレビュー機能の追加
        if st.session_state.get('user_info'):
            st.subheader("プレビュー")
            preview_html = preview_care_plan_table(st.session_state.user_info)
            st.markdown(preview_html, unsafe_allow_html=True)
        
        # 電子署名・印鑑設定
        signature_and_seal = upload_signature_and_seal()
        
        if st.button("基本情報を保存", type="primary"):
            if name and care_manager and care_office:
                st.session_state.user_info = {
                    "name": name,
                    "birth_date": birth_date.strftime('%Y年%m月%d日'),
                    "address": address,
                    "care_manager": care_manager,
                    "care_office": care_office,
                    "office_address": office_address,
                    "plan_date": plan_date.strftime('%Y年%m月%d日'),
                    "initial_plan_date": initial_plan_date.strftime('%Y年%m月%d日'),
                    "certification_date": certification_date.strftime('%Y年%m月%d日'),
                    "valid_from": valid_from.strftime('%Y年%m月%d日'),
                    "valid_to": valid_to.strftime('%Y年%m月%d日'),
                    "care_level": care_level,
                    "plan_status": {
                        "initial": initial_plan,
                        "introduced": introduced_plan,
                        "continuous": continuous_plan,
                        "certified": certified,
                        "applying": applying
                    },
                    "client_family_intentions": client_family_intentions,
                    "certification_opinion": certification_opinion,
                    "support_policy": support_policy,
                    "care_reason": care_reason + (f"（{other_reason}）" if care_reason == "3. その他" else ""),
                    "consent": {
                        "date": consent_date.strftime('%Y年%m月%d日'),
                        "signature": signature
                    },
                    "signature_and_seal": signature_and_seal
                }
                st.success("基本情報が保存されました")
                st.write(st.session_state.user_info)
            else:
                st.warning("必須項目（利用者名、計画作成者氏名、事業所名）を入力してください")

    elif page == "ADLデータ入力":
        st.header("ADLデータ入力")
        
        adl_categories = {
            "基本動作": ["食事", "排泄", "入浴", "移動", "着替え", "整容"],
            "認知・コミュニケーション": ["コミュニケーション", "認知機能", "睡眠"],
            "社会生活": ["服薬管理", "金銭管理", "買い物"]
        }
        
        tabs = st.tabs(list(adl_categories.keys()))
        all_adl_data = {}
        
        for tab, (category, items) in zip(tabs, adl_categories.items()):
            with tab:
                category_data = render_adl_input_section(items, category)
                all_adl_data.update(category_data)
        
        if st.button("ADLデータを保存", type="primary"):
            st.session_state.adl_data = all_adl_data
            st.success("ADLデータが保存されました")
            st.write(pd.DataFrame([all_adl_data]).T)
            
    elif page == "ケアプラン生成":
        st.header("ケアプラン生成")
        
        if 'user_info' not in st.session_state or 'adl_data' not in st.session_state:
            st.warning("基本情報とADLデータを先に入力してください")
            return
        
        # 生成済みケアプランの状態管理
        if 'current_care_plan' not in st.session_state:
            st.session_state.current_care_plan = None
            st.session_state.current_client_needs = None
        
        if not st.session_state.current_care_plan:
            st.subheader("利用者の要望")
            client_needs = st.text_area(
                "具体的な要望を入力してください",
                height=100,
                placeholder="例：母親の結婚式に参加したい、自宅で生活を続けたい、趣味の園芸を続けたい"
            )
            
            if st.button("ケアプランを生成", type="primary"):
                if not client_needs:
                    st.warning("利用者の要望を入力してください")
                    return
                
                with st.spinner("ケアプランを生成中..."):
                    care_plan = generate_care_plan(
                        st.session_state.user_info,
                        st.session_state.adl_data,
                        client_needs
                    )
                    
                    if care_plan:
                        st.session_state.current_care_plan = care_plan
                        st.session_state.current_client_needs = client_needs
                        st.rerun()
        
        # 生成済みケアプランの表示
        if st.session_state.current_care_plan:
            st.success("ケアプランが生成されました")
            
            # 新しいケアプランの生成ボタン
            if st.button("新しいケアプランを生成"):
                st.session_state.current_care_plan = None
                st.session_state.current_client_needs = None
                st.rerun()
            
            st.subheader("生成されたケアプラン")
            st.markdown(st.session_state.current_care_plan)
            
            # ダウンロードボタンのコンテナ
            download_container = st.container()
            with download_container:
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    # 辞書型を文字列に変換
                    care_plan_text = json.dumps(st.session_state.current_care_plan, ensure_ascii=False, indent=2)
                    st.download_button(
                        "テキスト形式でダウンロード",
                        care_plan_text,
                        "care_plan.txt",
                        "text/plain",
                        use_container_width=True
                    )
                
                with col2:
                    excel_buffer = create_care_plan_excel(
                        st.session_state.user_info,
                        st.session_state.adl_data,
                        st.session_state.current_care_plan
                    )
                    if excel_buffer:
                        st.download_button(
                            "エクセル形式でダウンロード",
                            excel_buffer,
                            "care_plan.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                
                with col3:
                    pdf_buffer = create_care_plan_pdf(
                        st.session_state.user_info,
                        st.session_state.adl_data,
                        st.session_state.current_care_plan
                    )
                    if pdf_buffer:
                        st.download_button(
                            "PDF形式でダウンロード",
                            pdf_buffer,
                            "care_plan.pdf",
                            "application/pdf",
                            use_container_width=True
                        )
                
                # テンプレートExcelへの出力ボタン
                st.button("テンプレートExcelに出力（outputフォルダ）", 
                          on_click=lambda: export_care_plan_to_excel_template(
                              st.session_state.user_info,
                              st.session_state.adl_data,
                              st.session_state.current_care_plan
                          ),
                          use_container_width=True)
            
            # 履歴への保存
            if st.session_state.current_care_plan not in [h['care_plan'] for h in st.session_state.care_plan_history]:
                history_entry = {
                    'timestamp': datetime.now(),
                    'user_info': st.session_state.user_info,
                    'adl_data': st.session_state.adl_data,
                    'client_needs': st.session_state.current_client_needs,
                    'care_plan': st.session_state.current_care_plan
                }
                st.session_state.care_plan_history.append(history_entry)
    
    elif page == "履歴管理":
        st.header("履歴管理")
        
        if not st.session_state.care_plan_history:
            st.info("まだケアプラン履歴がありません")
            return
        
        # 履歴の検索機能
        search_query = st.text_input("🔍 利用者名で検索", "")
        
        # 履歴の並び替え
        sort_order = st.radio(
            "並び替え",
            ["新しい順", "古い順"],
            horizontal=True
        )
        
        # 履歴のフィルタリングと並び替え
        filtered_history = st.session_state.care_plan_history.copy()
        if search_query:
            filtered_history = [
                h for h in filtered_history 
                if search_query.lower() in h['user_info']['name'].lower()
            ]
        
        if sort_order == "古い順":
            filtered_history.reverse()
        
        # 履歴の表示
        for i, history in enumerate(filtered_history):
            with st.expander(
                f"ケアプラン #{len(filtered_history) - i if sort_order == '新しい順' else i + 1} - "
                f"{history['timestamp'].strftime('%Y/%m/%d %H:%M')} "
                f"({history['user_info']['name']}様)"
            ):
                col1, col2 = st.columns([3, 1])
                
                with col1:
                    st.subheader("基本情報")
                    st.write(history['user_info'])
                    
                    st.subheader("ADLデータ")
                    st.write(pd.DataFrame([history['adl_data']]).T)
                    
                    st.subheader("利用者の要望")
                    st.write(history['client_needs'])
                    
                    st.subheader("生成されたケアプラン")
                    st.markdown(history['care_plan'])
                
                with col2:
                    st.markdown("### ダウンロード")
                    
                    # ダウンロード形式の選択
                    download_format = st.selectbox(
                        "形式を選択",
                        ["PDF形式", "エクセル形式", "テキスト形式", "すべての形式（ZIP）"],
                        key=f"format_{i}"
                    )
                    
                    timestamp = history['timestamp'].strftime('%Y%m%d_%H%M')
                    
                    if download_format == "PDF形式":
                        pdf_buffer = create_care_plan_pdf(
                            history['user_info'],
                            history['adl_data'],
                            history['care_plan']
                        )
                        if pdf_buffer:
                            st.download_button(
                                "📄 PDFをダウンロード",
                                pdf_buffer,
                                f"care_plan_{timestamp}.pdf",
                                "application/pdf"
                            )
                    
                    elif download_format == "エクセル形式":
                        excel_buffer = create_care_plan_excel(
                            history['user_info'],
                            history['adl_data'],
                            history['care_plan']
                        )
                        if excel_buffer:
                            st.download_button(
                                "📊 エクセルをダウンロード",
                                excel_buffer,
                                f"care_plan_{timestamp}.xlsx",
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    
                    elif download_format == "テキスト形式":
                        st.download_button(
                            "📝 テキストをダウンロード",
                            history['care_plan'],
                            f"care_plan_{timestamp}.txt",
                            "text/plain"
                        )
                    
                    else:  # すべての形式
                        zip_buffer = create_download_package(
                            history['user_info'],
                            history['adl_data'],
                            history['care_plan'],
                            timestamp
                        )
                        if zip_buffer:
                            st.download_button(
                                "📦 すべての形式をダウンロード",
                                zip_buffer,
                                f"care_plan_{timestamp}.zip",
                                "application/zip"
                            )

if __name__ == "__main__":
    main()