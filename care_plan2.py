import streamlit as st
import pandas as pd
from datetime import datetime
from openai import OpenAI
import os
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import tempfile
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import zipfile
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib.units import mm

# 環境変数の読み込み
load_dotenv()

# OpenAI クライアントの設定
client = OpenAI(
    api_key=os.getenv('OPENAI_API_KEY')
)

# ページ設定
st.set_page_config(
    page_title="EGAO-AI デモ",
    page_icon="👥",
    layout="wide"
)

# セッション状態の初期化
if 'generated_care_plan' not in st.session_state:
    st.session_state.generated_care_plan = None

if 'care_plan_history' not in st.session_state:
    st.session_state.care_plan_history = []

def get_adl_status_color(status):
    """ADL状態に応じたカラーコードを返す"""
    colors = {
        "要全介助": "#ff6b6b",  # 赤 - 最も介助が必要
        "一部介助": "#ffd93d",  # 黄 - 部分的な介助が必要
        "見守り": "#a3dc2e",    # 薄緑 - 自立に近い
        "自立": "#4CAF50"       # 緑 - 完全に自立
    }
    return colors.get(status, "#ffffff")

def get_adl_description(item, status):
    """ADL項目と状態に応じた説明文を返す"""
    descriptions = {
        "食事": {
            "要全介助": "食事の全過程で介助が必要（食事の準備から片付けまで、食べる動作すべてに介助が必要）",
            "一部介助": "食事の一部で介助が必要（食べ物を刻む、スプーンで掬うなどの補助が必要）",
            "見守り": "声かけ・見守りが必要（自力で食べられるが、誤嚥防止などの観察が必要）",
            "自立": "自力で食事が可能（準備から片付けまで完全に自立している）"
        },
        "排泄": {
            "要全介助": "排泄の全過程で介助が必要（トイレまでの移動、衣服の着脱、排泄後の処理すべてに介助が必要）",
            "一部介助": "排泄の一部で介助が必要（衣服の着脱の補助や、後始末の一部介助が必要）",
            "見守り": "声かけ・見守りが必要（自力で可能だが、安全確認のため見守りが必要）",
            "自立": "自力で排泄が可能（トイレまでの移動から後始末まで完全に自立している）"
        },
        "入浴": {
            "要全介助": "入浴の全過程で介助が必要（浴室への移動、衣服の着脱、洗体、洗髪すべてに介助が必要）",
            "一部介助": "入浴の一部で介助が必要（背中を洗う、髪を洗うなどの部分的な介助が必要）",
            "見守り": "声かけ・見守りが必要（自力で入浴可能だが、転倒防止のため見守りが必要）",
            "自立": "自力で入浴が可能（準備から後片付けまで完全に自立している）"
        },
        "移動": {
            "要全介助": "移動の全過程で介助が必要（ベッドから車椅子への移乗を含め、すべての移動に介助が必要）",
            "一部介助": "移動の一部で介助が必要（歩行器や杖を使用し、部分的な支援が必要）",
            "見守り": "声かけ・見守りが必要（自力で移動可能だが、安全確認のため見守りが必要）",
            "自立": "自力で移動が可能（補助具の使用の有無に関わらず、安全に移動できる）"
        },
        "着替え": {
            "要全介助": "着替えの全過程で介助が必要（衣服の選択から着脱まですべてに介助が必要）",
            "一部介助": "着替えの一部で介助が必要（ボタンの掛け外しや、靴下の着脱などの補助が必要）",
            "見守り": "声かけ・見守りが必要（自力で着替え可能だが、確認や声かけが必要）",
            "自立": "自力で着替えが可能（衣服の選択から着脱まで完全に自立している）"
        },
        "整容": {
            "要全介助": "整容の全過程で介助が必要（洗顔、歯磨き、整髪すべてに介助が必要）",
            "一部介助": "整容の一部で介助が必要（髭剃りや整髪などの部分的な介助が必要）",
            "見守り": "声かけ・見守りが必要（自力で整容可能だが、確認や声かけが必要）",
            "自立": "自力で整容が可能（身だしなみを整えることが完全に自立している）"
        },
        "コミュニケーション": {
            "要全介助": "コミュニケーションが極めて困難（意思疎通がほとんど取れない）",
            "一部介助": "基本的な意思疎通に支援が必要（簡単な言葉や身振りでの意思疎通が可能）",
            "見守り": "時々支援が必要（複雑な会話に困難があるが、基本的な意思疎通は可能）",
            "自立": "円滑なコミュニケーションが可能（会話に問題なく参加できる）"
        },
        "認知機能": {
            "要全介助": "重度の認知機能低下（日常生活のほとんどの判断に支援が必要）",
            "一部介助": "中等度の認知機能低下（日常生活の一部の判断に支援が必要）",
            "見守り": "軽度の認知機能低下（時々判断に迷いがあるが、声かけで対応可能）",
            "自立": "認知機能は良好（日常生活の判断に問題なし）"
        },
        "睡眠": {
            "要全介助": "睡眠のリズムが大きく乱れている（昼夜逆転や不眠が顕著）",
            "一部介助": "睡眠に問題があり支援が必要（入眠困難や中途覚醒がある）",
            "見守り": "時々睡眠に乱れがある（軽度の不眠や早朝覚醒がある）",
            "自立": "良好な睡眠が取れている（睡眠のリズムが安定している）"
        },
        "服薬管理": {
            "要全介助": "服薬管理が全くできない（薬の準備から服用まですべてに介助が必要）",
            "一部介助": "服薬管理に部分的な支援が必要（薬の仕分けや時間の管理に援助が必要）",
            "見守り": "声かけがあれば服薬管理可能（確認や促しが必要）",
            "自立": "自力で服薬管理が可能（処方薬の管理を適切に行える）"
        },
        "金銭管理": {
            "要全介助": "金銭管理が全くできない（収支の理解や管理がまったくできない）",
            "一部介助": "基本的な金銭管理に支援が必要（日々の出納の一部に支援が必要）",
            "見守り": "声かけがあれば金銭管理可能（確認や助言が必要）",
            "自立": "自力で金銭管理が可能（収支の把握や管理が適切にできる）"
        },
        "買い物": {
            "要全介助": "買い物が全くできない（商品の選択から支払いまですべてに介助が必要）",
            "一部介助": "買い物に部分的な支援が必要（商品の選択や支払いの一部に援助が必要）",
            "見守り": "声かけがあれば買い物可能（確認や助言が必要）",
            "自立": "自力で買い物が可能（必要な物品の購入を適切に行える）"
        }
    }
    default_descriptions = {
        "要全介助": "常時介助が必要（自力での実施が困難）",
        "一部介助": "部分的な介助が必要（一部自力で行えるが支援が必要）",
        "見守り": "声かけ・見守りが必要（安全確認のため観察が必要）",
        "自立": "自力で可能（支援なしで実施できる）"
    }
    return descriptions.get(item, default_descriptions).get(status, "")

def create_care_plan_excel(user_info, adl_data, care_plan):
    """ケアプランをcare_plan_1.xlsの形式で生成"""
    try:
        wb = openpyxl.Workbook()
        
        # 第1表シート
        ws1 = wb.active
        ws1.title = "第1表"
        
        # ヘッダー部分
        ws1["A1"] = "居宅サービス計画書（1）"
        ws1["A2"] = f"作成年月日：{datetime.now().strftime('%Y年%m月%d日')}"
        
        # 基本情報
        ws1["A4"] = "利用者基本情報"
        ws1["A5"] = f"氏名：{user_info['name']}"
        ws1["C5"] = f"性別：{user_info['gender']}"
        ws1["E5"] = f"年齢：{user_info['age']}歳"
        ws1["A6"] = f"要介護度：{user_info['care_level']}"
        ws1["C6"] = f"認定日：{datetime.now().strftime('%Y年%m月%d日')}"
        ws1["A7"] = f"家族構成：{user_info['family_structure']}"
        ws1["A8"] = f"キーパーソン：{user_info['key_person']}"
        
        # ADL評価
        ws1["A10"] = "ADL評価"
        row = 11
        for item, status in adl_data.items():
            ws1[f"A{row}"] = item
            ws1[f"B{row}"] = status
            row += 1
        
        # ケアプラン内容
        sections = care_plan.split("【")
        current_sheet = ws1
        current_row = row + 2
        
        for section in sections:
            if not section.strip():
                continue
            
            if "第2表" in section:
                current_sheet = wb.create_sheet("第2表")
                current_row = 1
            elif "第3表" in section:
                current_sheet = wb.create_sheet("第3表")
                current_row = 1
            
            lines = section.split("\n")
            for line in lines:
                if line.strip():
                    current_sheet[f"A{current_row}"] = line
                    current_row += 1
        
        # スタイル設定
        for ws in wb.worksheets:
            # 列幅の設定
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 25
            
            # フォントとセル書式の設定
            for row in ws.rows:
                for cell in row:
                    cell.font = openpyxl.styles.Font(name='游ゴシック', size=10)
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
        
        # Excelファイルをバイトストリームとして保存
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer

    except Exception as e:
        st.error(f"エクセルファイル生成中にエラーが発生しました: {str(e)}")
        return None

def generate_care_plan(user_info, adl_data, client_needs):
    """OpenAI APIを使用してケアプラン生成"""
    try:
        # client_informationの形式に合わせたプロンプト構築
        prompt = f"""
以下の介護者情報を元に、居住サービス計画書（第1表～第3表）を作成してください。

【利用者基本情報】
氏名: {user_info['name']}
性別: {user_info['gender']}
年齢: {user_info['age']}歳
要介護度: {user_info['care_level']}
家族構成: {user_info['family_structure']}
キーパーソン: {user_info['key_person']}

【ADL評価】
{pd.DataFrame([adl_data]).T.to_string()}

【利用者・家族の要望】
{client_needs}

以下の形式で出力してください：

【第1表】
■利用者・家族の意向と総合的な援助の方針
■解決すべき課題
■サービス提供の意向

【第2表】
■生活全般の解決すべき課題
■長期目標（6ヶ月）
■短期目標（3ヶ月）
■サービス内容と種別
■担当者と頻度

【第3表】
■週間サービス計画
■主な日常生活上の活動
■家族の支援・連携内容
■サービス提供上の留意事項
"""

        response = client.chat.completions.create(
            model="gpt-4-turbo-preview",
            messages=[
                {"role": "system", "content": "あなたは経験豊富な介護支援専門員です。"},
                {"role": "user", "content": prompt}
            ],
            temperature=0.7,
            max_tokens=2000
        )

        return response.choices[0].message.content

    except Exception as e:
        st.error(f"ケアプラン生成中にエラーが発生しました: {str(e)}")
        return None

def render_adl_input_section(items, category_name):
    """ADL入力セクションのレンダリング"""
    st.markdown(f"### {category_name}", unsafe_allow_html=True)
    category_data = {}
    
    for item in items:
        st.markdown("""
            <div style='
                border: 1px solid #e0e0e0;
                border-radius: 10px;
                padding: 15px;
                margin: 10px 0;
                background-color: #ffffff;
                box-shadow: 0 2px 4px rgba(0,0,0,0.05);
            '>
        """, unsafe_allow_html=True)
        
        st.markdown(f"#### {item}")
        status = st.selectbox(
            "状態を選択",
            ["要全介助", "一部介助", "見守り", "自立"],
            key=f"adl_{item}"
        )
        category_data[item] = status
        
        description = get_adl_description(item, status)
        if description:
            st.markdown(
                f"""<div style='
                    padding: 10px;
                    border-radius: 5px;
                    background-color: {get_adl_status_color(status)}15;
                    border-left: 5px solid {get_adl_status_color(status)};
                    margin-top: 10px;
                '>{description}</div>""",
                unsafe_allow_html=True
            )
        
        st.markdown("</div>", unsafe_allow_html=True)
    
    return category_data

def create_care_plan_pdf(user_info, adl_data, care_plan):
    """ケアプランをPDFファイルとして生成"""
    try:
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=A4,
            title="ケアプラン",
            rightMargin=20*mm,
            leftMargin=20*mm,
            topMargin=20*mm,
            bottomMargin=20*mm
        )
        
        # 日本語フォントの登録
        pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
        
        # スタイル設定
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            fontName='HeiseiKakuGo-W5',
            spaceAfter=30
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            fontName='HeiseiKakuGo-W5',
            spaceAfter=20
        )
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            fontName='HeiseiKakuGo-W5',
            leading=14
        )
        
        # ドキュメント要素
        elements = []
        
        elements.append(Paragraph("居宅サービス計画書", title_style))
        elements.append(Spacer(1, 20))
        
        elements.append(Paragraph("利用者基本情報", heading_style))
        for key, value in user_info.items():
            elements.append(Paragraph(f"{key}: {value}", normal_style))
        elements.append(Spacer(1, 20))
        
        elements.append(Paragraph("ADL評価", heading_style))
        for item, status in adl_data.items():
            elements.append(Paragraph(f"{item}: {status}", normal_style))
        elements.append(Spacer(1, 20))
        
        elements.append(Paragraph("ケアプラン内容", heading_style))
        for section in care_plan.split('\n\n'):
            elements.append(Paragraph(section, normal_style))
            elements.append(Spacer(1, 10))
        
        doc.build(elements)
        pdf_buffer.seek(0)
        return pdf_buffer
        
    except Exception as e:
        st.error(f"PDF生成中にエラーが発生しました: {str(e)}")
        return None

def create_download_package(user_info, adl_data, care_plan, timestamp):
    """各形式のファイルを含むZIPパッケージを作成"""
    try:
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            # テキストファイル
            zf.writestr(
                f"care_plan_{timestamp}.txt",
                care_plan
            )
            
            # Excelファイル
            excel_buffer = create_care_plan_excel(user_info, adl_data, care_plan)
            if excel_buffer:
                zf.writestr(
                    f"care_plan_{timestamp}.xlsx",
                    excel_buffer.getvalue()
                )
            
            # PDFファイル
            pdf_buffer = create_care_plan_pdf(user_info, adl_data, care_plan)
            if pdf_buffer:
                zf.writestr(
                    f"care_plan_{timestamp}.pdf",
                    pdf_buffer.getvalue()
                )
        
        zip_buffer.seek(0)
        return zip_buffer
    
    except Exception as e:
        st.error(f"パッケージ作成中にエラーが発生しました: {str(e)}")
        return None

def main():
    st.markdown("""
        <h1 style='color: #1E88E5; font-size: 32px;'>
            EGAO-AI ケアプラン作成支援システム
        </h1>
    """, unsafe_allow_html=True)
    
    # サイドバーのスタイリング
    with st.sidebar:
        st.markdown("""
            <h2 style='color: #333; font-size: 24px; font-weight: 600;'>
                メニュー
            </h2>
        """, unsafe_allow_html=True)
        page = st.radio(
            "選択してください",
            ["基本情報入力", "ADLデータ入力", "ケアプラン生成", "履歴管理"]
        )

    # ADLカテゴリーの定義を更新
    adl_categories = {
        "🚶 基本動作": ["食事", "排泄", "入浴", "移動", "着替え", "整容"],
        "🧠 認知・コミュニケーション": ["コミュニケーション", "認知機能", "睡眠"],
        "🏠 社会生活": ["服薬管理", "金銭管理", "買い物"]
    }

    # 各ページのヘッダースタイリング
    if page == "基本情報入力":
        st.markdown("""
            <h2 style='color: #333; font-size: 28px; font-weight: 600; margin-bottom: 20px;'>
                基本情報入力
            </h2>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        with col1:
            name = st.text_input("利用者名（被介護者の氏名）")
            age = st.number_input("年齢（被介護者の年齢）", min_value=0, max_value=150)
            gender = st.selectbox("性別（被介護者の性別）", ["選択してください", "男性", "女性"])
            
        with col2:
            care_level = st.selectbox(
                "要介護度（介護保険で認定された介護度）",
                ["選択してください", "要介護1", "要介護2", "要介護3", "要介護4", "要介護5"]
            )
            family_structure = st.text_input("家族構成（同居家族や主な支援者の状況）", 
                                          placeholder="例：長男夫婦と同居、配偶者は他界")
            key_person = st.text_input("キーパーソン（主な介護者や連絡先となる方）",
                                     placeholder="例：長男（同居）、次女（近所に在住）")
        
        if st.button("基本情報を保存", type="primary"):
            if name and age > 0 and gender != "選択してください" and care_level != "選択してください":
                st.session_state.user_info = {
                    "name": name,
                    "age": age,
                    "gender": gender,
                    "care_level": care_level,
                    "family_structure": family_structure,
                    "key_person": key_person
                }
                st.success("基本情報が保存されました")
                st.write(st.session_state.user_info)
            else:
                st.warning("必須項目を入力してください")

    elif page == "ADLデータ入力":
        st.header("ADLデータ入力")
        
        adl_categories = {
            "基本動作": ["食事", "排泄", "入浴", "移動", "着替え", "整容"],
            "認知・コミュニケーション": ["コミュニケーション", "認知機能", "睡眠"],
            "社会生活": ["服薬管理", "金銭管理", "買い物"]
        }
        
        tabs = st.tabs(list(adl_categories.keys()))
        all_adl_data = {}
        
        for tab, (category, items) in zip(tabs, adl_categories.items()):
            with tab:
                category_data = render_adl_input_section(items, category)
                all_adl_data.update(category_data)
        
        if st.button("ADLデータを保存", type="primary"):
            st.session_state.adl_data = all_adl_data
            st.success("ADLデータが保存されました")
            st.write(pd.DataFrame([all_adl_data]).T)
            
    elif page == "ケアプラン生成":
        st.header("ケアプラン生成")
        
        if 'user_info' not in st.session_state or 'adl_data' not in st.session_state:
            st.warning("基本情報とADLデータを先に入力してください")
            return
        
        st.subheader("利用者の要望")
        client_needs = st.text_area(
            "具体的な要望を入力してください",
            height=100,
            placeholder="例：母親の結婚式に参加したい、自宅で生活を続けたい、趣味の園芸を続けたい"
        )
        
        if st.button("ケアプランを生成", type="primary"):
            if not client_needs:
                st.warning("利用者の要望を入力してください")
                return
                
            with st.spinner("ケアプランを生成中..."):
                care_plan = generate_care_plan(
                    st.session_state.user_info,
                    st.session_state.adl_data,
                    client_needs
                )
                
                if care_plan:
                    st.session_state.generated_care_plan = care_plan
                    
                    # 履歴に保存
                    history_entry = {
                        'timestamp': datetime.now(),
                        'user_info': st.session_state.user_info,
                        'adl_data': st.session_state.adl_data,
                        'client_needs': client_needs,
                        'care_plan': care_plan
                    }
                    st.session_state.care_plan_history.append(history_entry)
                    
                    st.success("ケアプランが生成されました")
                    
                    st.subheader("生成されたケアプラン")
                    st.markdown(care_plan)
                    
                    # ダウンロードボタンのセクション
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.download_button(
                            "テキスト形式でダウンロード",
                            care_plan,
                            "care_plan.txt",
                            "text/plain"
                        )
                    
                    with col2:
                        excel_buffer = create_care_plan_excel(
                            st.session_state.user_info,
                            st.session_state.adl_data,
                            care_plan
                        )
                        if excel_buffer:
                            st.download_button(
                                "エクセル形式でダウンロード",
                                excel_buffer,
                                "care_plan.xlsx",
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    
                    with col3:
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
                        zip_buffer = create_download_package(
                            st.session_state.user_info,
                            st.session_state.adl_data,
                            care_plan,
                            timestamp
                        )
                        if zip_buffer:
                            st.download_button(
                                "すべての形式をダウンロード",
                                zip_buffer,
                                f"care_plan_{timestamp}.zip",
                                "application/zip"
                            )
    
    elif page == "履歴管理":
        st.header("履歴管理")
        
        if not st.session_state.care_plan_history:
            st.info("まだケアプラン履歴がありません")
            return
        
        # 履歴の検索機能
        search_query = st.text_input("🔍 利用者名で検索", "")
        
        # 履歴の並び替え
        sort_order = st.radio(
            "並び替え",
            ["新しい順", "古い順"],
            horizontal=True
        )
        
        # 履歴のフィルタリングと並び替え
        filtered_history = st.session_state.care_plan_history.copy()
        if search_query:
            filtered_history = [
                h for h in filtered_history 
                if search_query.lower() in h['user_info']['name'].lower()
            ]
        
        if sort_order == "古い順":
            filtered_history.reverse()
        
        # 履歴の表示
        for i, history in enumerate(filtered_history):
            with st.expander(
                f"ケアプラン #{len(filtered_history) - i if sort_order == '新しい順' else i + 1} - "
                f"{history['timestamp'].strftime('%Y/%m/%d %H:%M')} "
                f"({history['user_info']['name']}様)"
            ):
                col1, col2 = st.columns([3, 1])
                
                with col1:
                    st.subheader("基本情報")
                    st.write(history['user_info'])
                    
                    st.subheader("ADLデータ")
                    st.write(pd.DataFrame([history['adl_data']]).T)
                    
                    st.subheader("利用者の要望")
                    st.write(history['client_needs'])
                    
                    st.subheader("生成されたケアプラン")
                    st.markdown(history['care_plan'])
                
                with col2:
                    st.markdown("### ダウンロード")
                    
                    # ダウンロード形式の選択
                    download_format = st.selectbox(
                        "形式を選択",
                        ["PDF形式", "エクセル形式", "テキスト形式", "すべての形式（ZIP）"],
                        key=f"format_{i}"
                    )
                    
                    timestamp = history['timestamp'].strftime('%Y%m%d_%H%M')
                    
                    if download_format == "PDF形式":
                        pdf_buffer = create_care_plan_pdf(
                            history['user_info'],
                            history['adl_data'],
                            history['care_plan']
                        )
                        if pdf_buffer:
                            st.download_button(
                                "📄 PDFをダウンロード",
                                pdf_buffer,
                                f"care_plan_{timestamp}.pdf",
                                "application/pdf"
                            )
                    
                    elif download_format == "エクセル形式":
                        excel_buffer = create_care_plan_excel(
                            history['user_info'],
                            history['adl_data'],
                            history['care_plan']
                        )
                        if excel_buffer:
                            st.download_button(
                                "📊 エクセルをダウンロード",
                                excel_buffer,
                                f"care_plan_{timestamp}.xlsx",
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    
                    elif download_format == "テキスト形式":
                        st.download_button(
                            "📝 テキストをダウンロード",
                            history['care_plan'],
                            f"care_plan_{timestamp}.txt",
                            "text/plain"
                        )
                    
                    else:  # すべての形式
                        zip_buffer = create_download_package(
                            history['user_info'],
                            history['adl_data'],
                            history['care_plan'],
                            timestamp
                        )
                        if zip_buffer:
                            st.download_button(
                                "📦 すべての形式をダウンロード",
                                zip_buffer,
                                f"care_plan_{timestamp}.zip",
                                "application/zip"
                            )

if __name__ == "__main__":
    main()